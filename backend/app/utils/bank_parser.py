"""
Çoklu banka ekstresi ayrıştırıcı.

Desteklenen bankalar:
  - Ziraat Bankası (PDF tablo)
  - TEB (PDF tablo — 10 kolon)
  - Garanti BBVA (PDF tablo — 5 kolon, +/- ön ek, döviz son ek)
  - Yapı Kredi (PDF metin — DD/MM/YYYY, döviz son ek)
  - VakıfBank (PDF metin — çok satırlı açıklama, Tarih Saat İşlemNo Miktar Bakiye İşlemAdı)
  - Halkbank (Excel — header satırları 0-22, veri satırı 23+)
  - QNB Finansbank (PDF tablo — İngilizce sayı formatı: 1,000.00)
"""
import hashlib
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import List, Optional

from app.utils.bank_parse_helpers import (
    _detect_number_format,
    _extract_trailing_numbers,
    _normalize_tr,
    _smart_parse_number,
    _strip_currency_suffix,
    parse_date_tr,
    parse_english_number,
    parse_turkish_number,
)


@dataclass
class ParsedHeader:
    iban: Optional[str] = None
    currency: Optional[str] = None
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    branch_name: Optional[str] = None
    account_no: Optional[str] = None
    holder_name: Optional[str] = None


@dataclass
class ParsedTransaction:
    date: date
    receipt_no: Optional[str]
    description: str
    amount: float  # imzalı: negatif = borç, pozitif = alacak
    balance: Optional[float]
    type: str  # income / expense
    tx_hash: str = ""
    time: Optional[str] = None  # HH:MM formatı — aynı gün sıralama için


@dataclass
class ParseResult:
    header: ParsedHeader = field(default_factory=ParsedHeader)
    transactions: List[ParsedTransaction] = field(default_factory=list)


def _balance_chain_score(txs: List[ParsedTransaction]) -> int:
    """Bakiye zincirinin tutarlılık puanı (yüksek = doğru sıra).

    tx[i].balance ≈ tx[i-1].balance + tx[i].amount ise +1 puan.
    """
    score = 0
    for i in range(1, len(txs)):
        prev_bal = txs[i - 1].balance
        cur_bal = txs[i].balance
        cur_amt = txs[i].amount
        if prev_bal is not None and cur_bal is not None and cur_amt is not None:
            expected = prev_bal + cur_amt
            if abs(expected - cur_bal) < 0.02:
                score += 1
    return score


def _ensure_chronological_order(transactions: List[ParsedTransaction]) -> None:
    """İşlemleri kronolojik sıraya getir (eskiden yeniye).

    1) Tarih sırasına bakar — ilk tarih > son tarih ise listeyi çevirir.
    2) Aynı gün işlemlerinde bakiye zincirini kontrol eder —
       ters sıra daha tutarlıysa listeyi çevirir.
    3) Bakiye zinciri berabere ise saat veya dekont numarası ile karar verir.
    4) Hâlâ berabereyse ve tam 2 işlem varsa (tutarları birebir zıt — ör. kredi
       kullandırımı + aynı tutarın tam transferi): gelir gidere ÖNCE gelir varsayımı
       ile karar verilir (para harcanmadan önce alınmış olmalı). Bu durumda ileri/geri
       bakiye-zinciri skoru matematiksel olarak HER ZAMAN eşittir (amt[0]=-amt[1] iken
       ayırt edilemez) — canlı bulgu: Eximbank TRY ekstresi, aynı gün kredi kullandırımı
       + tam transfer, dosya sırası ters geldiğinde bakiye zinciri bunu yakalayamadı
       (2026-07-06 düzeltmesi).
    """
    if len(transactions) < 2:
        return

    first_date = transactions[0].date
    last_date = transactions[-1].date
    if first_date > last_date:
        transactions.reverse()
        return

    # Tarih aynı veya karışık → bakiye zinciriyle karar ver
    fwd_score = _balance_chain_score(transactions)
    rev_score = _balance_chain_score(list(reversed(transactions)))
    if rev_score > fwd_score:
        transactions.reverse()
        return

    # Bakiye zinciri berabere → saat veya dekont numarası ile tiebreak
    if fwd_score == rev_score:
        # Saat bilgisi varsa kullan (TEB SAAT kolonu)
        times = [tx.time for tx in transactions if tx.time]
        if len(times) >= 2:
            first_time = transactions[0].time
            last_time = transactions[-1].time
            if first_time and last_time and first_time > last_time:
                transactions.reverse()
                return

        # Saat yoksa dekont numarası sırası ile karar ver
        receipts = []
        for tx in transactions:
            if tx.receipt_no and tx.receipt_no.strip().isdigit():
                receipts.append(int(tx.receipt_no.strip()))
        if len(receipts) >= 2 and receipts[0] > receipts[-1]:
            transactions.reverse()
            return

        # Saat/dekont da yoksa (veya kararsızsa): tam 2 işlem + birebir zıt tutar
        # (bakiye zinciri matematiksel olarak ayırt edemez) → gelir gidere ÖNCE gelir
        if len(transactions) == 2:
            a, b = transactions[0], transactions[1]
            if a.type != b.type and abs(abs(a.amount) - abs(b.amount)) < 0.02:
                if a.type == "expense" and b.type == "income":
                    transactions.reverse()
                return


def compute_tx_hash(tx_date: date, receipt_no: Optional[str], amount: float, description: str, seq: int = 0) -> str:
    """Mükerrer tespit için SHA-256 hash hesapla.
    seq parametresi aynı gün+tutar+açıklama olan farklı işlemleri ayırt eder.
    """
    desc_key = (description or "")[:50].strip()
    raw = f"{tx_date.isoformat()}|{receipt_no or ''}|{amount:.2f}|{desc_key}"
    if seq > 0:
        raw += f"|{seq}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# ─── PDF Parser ──────────────────────────────────────────

TESSDATA_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "tessdata"
)


def _ocr_page(page) -> str:
    """Taranmış PDF sayfasından OCR ile metin çıkar."""
    import subprocess
    import tempfile

    try:
        img = page.to_image(resolution=300)
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            img.save(tmp.name)
            tmp_path = tmp.name

        script = f"""
import os, sys
os.environ['TESSDATA_PREFIX'] = '{TESSDATA_DIR}'
import tesserocr
from PIL import Image
img = Image.open('{tmp_path}')
with tesserocr.PyTessBaseAPI(path='{TESSDATA_DIR}', lang='tur+eng') as api:
    api.SetPageSegMode(tesserocr.PSM.SINGLE_BLOCK)
    api.SetImage(img)
    text = api.GetUTF8Text()
sys.stdout.write(text)
"""
        result = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True, text=True, timeout=60,
        )
        os.unlink(tmp_path)

        if result.returncode == 0:
            return result.stdout
        return ""
    except Exception:
        return ""


def parse_pdf(file_path: str) -> ParseResult:
    """PDF banka ekstresi ayrıştır (tüm bankalar)."""
    import pdfplumber

    result = ParseResult()

    with pdfplumber.open(file_path) as pdf:
        full_text = ""
        all_tables = []

        for pg in pdf.pages:
            page_text = pg.extract_text() or ""
            full_text += page_text + "\n"

            tables = pg.extract_tables()
            for table in tables:
                all_tables.extend(table)

        # Metin veya tablo bulunamadıysa OCR dene
        # Taranmış PDF'lerde pdfplumber birkaç karakter çıkarabilir,
        # bu yüzden 100 karakterden kısa metinleri de "boş" sayıyoruz
        text_too_short = len(full_text.strip()) < 100
        if (text_too_short and not all_tables):
            for pg in pdf.pages:
                ocr_text = _ocr_page(pg)
                if ocr_text:
                    full_text += ocr_text + "\n"

        # Header bilgilerini çıkar
        result.header = _extract_header(full_text)

        # VakıfBank formatını tespit et (TARİH SAAT İŞLEM NO MİKTAR BAKİYE İŞLEM ADI)
        is_vakifbank = bool(re.search(
            r"TAR[İI]H\s+SAAT\s+[İI][ŞS]LEM\s+NO\s+M[İI]KTAR\s+BAK[İI]YE",
            full_text, re.IGNORECASE
        ))

        if is_vakifbank:
            result.transactions = _parse_vakifbank_text(full_text)
        else:
            # İşlemleri tablolardan çıkar (akıllı kolon tespiti)
            if all_tables:
                result.transactions = _parse_table_rows_smart(all_tables)

            # Tablo bulunamadıysa veya sonuç yoksa metin tabanlı dene
            if not result.transactions:
                result.transactions = _parse_text_rows(full_text)

    # Kronolojik sıraya getir (eskiden yeniye)
    _ensure_chronological_order(result.transactions)

    # NOT: _repair_balance_gaps() devre dışı bırakıldı.
    # Birden fazla işlem eksik olduğunda tek bir hatalı kayıt oluşturuyor
    # ve DB'deki ID sıralamasıyla (ORDER BY date, id) bakiye zinciri bozuluyor.

    return result


def _extract_header(text: str) -> ParsedHeader:
    """PDF metninden header bilgilerini çıkar."""
    header = ParsedHeader()
    flat = re.sub(r"\s+", " ", text)

    # ─── IBAN — BBAN alfanümerik olabilir (Halkbank harf içeren hesap no) ───
    digits_only = re.sub(r"(?<=[A-Z0-9])\s+(?=[A-Z0-9])", "", flat, flags=re.IGNORECASE)
    m = re.search(r"IBAN\s*:?\s*\d?\s*(TR\d{2}[A-Z0-9]{22})", digits_only, re.IGNORECASE)
    if not m:
        iban_pos = digits_only.upper().find("IBAN")
        if iban_pos >= 0:
            m = re.search(r"(TR\d{2}[A-Z0-9]{22})", digits_only[iban_pos:iban_pos + 80], re.IGNORECASE)
    if not m:
        m = re.search(r"(TR\d{2}[A-Z0-9]{22})", digits_only, re.IGNORECASE)
    if m:
        header.iban = m.group(1).upper()

    # ─── Döviz Cinsi ───
    flat_norm = _normalize_tr(flat)
    # Açık etiketler
    m = re.search(r"doviz\s+cinsi\s*:?\s*(TRY|TL|EUR|EURO|USD|DOLAR)\b", flat_norm, re.IGNORECASE)
    if not m:
        dov_pos = re.search(r"doviz\s+cinsi", flat_norm, re.IGNORECASE)
        if dov_pos:
            m = re.search(r":?\s*(TRY|TL|EUR|EURO|USD|DOLAR)\b", flat_norm[dov_pos.end():], re.IGNORECASE)
    # "Vadesiz TL/EUR/USD" kalıbı (Yapı Kredi, VakıfBank, Garanti)
    if not m:
        m = re.search(r"vadesiz\s+(TL|TRY|EUR|USD)\b", flat_norm, re.IGNORECASE)
    # "TL/EUR/USD Vadesiz" kalıbı (QNB Finansbank — ters sıra)
    if not m:
        m = re.search(r"(TL|TRY|EUR|USD)\s+vadesiz", flat_norm, re.IGNORECASE)
    # "Bakiye : 1.282,20 TL/EUR/USD" kalıbı (Garanti BBVA)
    if not m:
        m = re.search(r"bakiye\s*:\s*[\d.,]+\s+(TL|TRY|EUR|USD)\b", flat_norm, re.IGNORECASE)
    if m:
        currency = m.group(1).upper()
        if currency in ("TL", "TRY", "TÜRK"):
            header.currency = "TRY"
        elif currency in ("EUR", "EURO"):
            header.currency = "EUR"
        elif currency in ("USD", "DOLAR"):
            header.currency = "USD"
        else:
            header.currency = currency[:3]

    # ─── Dönem ───
    m = re.search(r"D[öo]n?em\s*:?\s*(\d{2}[./]\d{2}[./]\d{4})\s*[-–]\s*(\d{2}[./]\d{2}[./]\d{4})", flat)
    if not m:
        m = re.search(r"(\d{2}[./]\d{2}[./]\d{4})\s*[-–]\s*(\d{2}[./]\d{2}[./]\d{4})", flat)
    if m:
        header.period_start = parse_date_tr(m.group(1))
        header.period_end = parse_date_tr(m.group(2))

    # ─── Şube ───
    m = re.search(r"[SŞ]ube\s+(?:Kodu|Ad[ıi])\s*:?\s*(.+?)(?:\n|$)", text)
    if not m:
        m = re.search(r"[SŞ]ube\s+Kodu\s*:?\s*(.+?)(?:M[üu]|IBAN|D[öo]v)", flat)
    # QNB format: "Şube : DEMOKR*** BULV***"
    if not m:
        m = re.search(r"[SŞ]ube\s*:\s*(.+?)(?:\n|Tarih|$)", text)
    if m:
        header.branch_name = m.group(1).strip()

    # ─── Hesap No ───
    m = re.search(r"M[üu][sş]teri\s*/?Hesap\s*(?:No|Mo)\s*:?\s*(\S+)", flat)
    if m:
        header.account_no = m.group(1).strip()

    # ─── Hesap Sahibi ───
    # Garanti format: "Sayın XXX, DD Ay YYYY tarihi saat ..."
    m = re.search(r"Say[ıi]n\s+(.+?),\s*\d{2}\s+", flat)
    if not m:
        # Ziraat/VakıfBank format: "Sayın : XXX\n" veya "Sayın XXX\nAdres"
        m = re.search(r"Say[ıi]n\s*:?\s*(.+?)(?:\n|Adres|$)", text)
    if not m:
        m = re.search(r"Say[ıi]n\s*:?\s*(.+?)(?:Adres|[SŞ]ube)", flat)
    # QNB format: "Unvan : MURAT-A TURİZM..."
    if not m:
        m = re.search(r"Unvan\s*:\s*(.+?)(?:\s+Hesap|\s+Vkn|\n|$)", flat)
    # VakıfBank: İlk satırda şirket adı
    if not m:
        m = re.search(r"Hesap\s+Hareketleri\s*\n(.+?)(?:\n|Hesap\s+No)", text)
    if m:
        name = re.sub(r"\s+", " ", m.group(1)).strip(" .:,")
        if name and 3 < len(name) < 150:
            header.holder_name = name

    return header


# ─── Akıllı Tablo Ayrıştırıcı ───────────────────────────

def _parse_table_rows_smart(rows: list) -> List[ParsedTransaction]:
    """Tablo satırlarını akıllı kolon tespitiyle ayrıştır.

    Desteklenen formatlar:
      - Ziraat: Tarih | Fiş No | Açıklama | Tutar | Bakiye
      - TEB: TARİH | VALÖR | SAAT | AÇIKLAMA | İŞLEM ÖZEL AÇIKLAMASI | İŞLEM TİPİ | TUTAR | BAKİYE | DEKONT NO | ALICI
      - Garanti: Tarih | Açıklama | Etiket | Tutar | Bakiye
      - QNB: İşlem Tarihi | Kanal | İşlem Açıklaması | Tutar | Bakiye (İngilizce sayı formatı)
    """
    transactions = []
    col_map = {}
    header_found = False
    data_rows_cells = []

    for row in rows:
        if not row:
            continue

        cells = [str(c).strip() if c else "" for c in row]
        joined = _normalize_tr(" ".join(cells))

        # Header satırını tespit et ve kolon haritasını oluştur
        if not header_found and "tarih" in joined:
            col_map = _detect_columns(cells)
            if col_map.get("date") is not None:
                header_found = True
            continue

        # Toplam/footer satırlarını atla
        if any(t in joined for t in ("toplam borc", "toplam alacak", "toplam", "sayfa")):
            continue

        if not header_found:
            continue

        data_rows_cells.append(cells)

    if not data_rows_cells:
        return transactions

    # Sayı formatını tespit et (QNB: English, diğerleri: Turkish)
    num_fmt = _detect_number_format(
        data_rows_cells, col_map.get("amount"), col_map.get("balance"),
    )
    number_parser = parse_english_number if num_fmt == "english" else parse_turkish_number

    for cells in data_rows_cells:
        tx = _try_parse_mapped_row(cells, col_map, number_parser)
        if tx:
            transactions.append(tx)

    return transactions


def _detect_columns(header_cells: list) -> dict:
    """Header hücre isimlerinden kolon indekslerini tespit et."""
    col_map = {}
    for i, h in enumerate(header_cells):
        hl = _normalize_tr(h)
        # Tarih (ilk bulunan "tarih")
        if "tarih" in hl and "date" not in col_map:
            col_map["date"] = i
        # Açıklama
        elif ("aciklama" in hl) and "desc" not in col_map:
            col_map["desc"] = i
        # Tutar / Miktar / İşlem Tutarı
        elif ("tutar" in hl or "miktar" in hl) and "amount" not in col_map:
            col_map["amount"] = i
        # Bakiye / Yeni Bakiye
        elif "bakiye" in hl and "balance" not in col_map:
            col_map["balance"] = i
        # Fiş No / Dekont No / İşlem No
        elif ("fis" in hl or "dekont" in hl or "islem no" in hl):
            col_map["receipt"] = i
        # Etiket (Garanti)
        elif "etiket" in hl:
            col_map["tag"] = i
        # Saat (TEB)
        elif "saat" in hl and "time" not in col_map:
            col_map["time"] = i
        # Valör — atla
    return col_map


def _try_parse_mapped_row(cells: list, col_map: dict, number_parser=None) -> Optional[ParsedTransaction]:
    """Kolon haritasına göre tek bir satırı ayrıştır."""
    if number_parser is None:
        number_parser = parse_turkish_number

    date_idx = col_map.get("date")
    if date_idx is None or date_idx >= len(cells):
        return None

    tx_date = parse_date_tr(cells[date_idx])
    if not tx_date:
        return None

    # Açıklama
    desc_idx = col_map.get("desc")
    description = cells[desc_idx].strip() if desc_idx is not None and desc_idx < len(cells) else ""

    # Etiket varsa açıklamaya ekle (Garanti)
    tag_idx = col_map.get("tag")
    if tag_idx is not None and tag_idx < len(cells) and cells[tag_idx].strip():
        tag = cells[tag_idx].strip()
        if tag.lower() not in ("diğer", "diger"):
            description = f"[{tag}] {description}"

    # Fiş No
    receipt_idx = col_map.get("receipt")
    receipt_no = None
    if receipt_idx is not None and receipt_idx < len(cells) and cells[receipt_idx].strip():
        receipt_no = cells[receipt_idx].strip()

    # Tutar
    amount_idx = col_map.get("amount")
    if amount_idx is None or amount_idx >= len(cells):
        return None
    amount = number_parser(cells[amount_idx])
    if amount is None:
        return None

    # Bakiye
    balance = None
    balance_idx = col_map.get("balance")
    if balance_idx is not None and balance_idx < len(cells):
        balance = number_parser(cells[balance_idx])

    # Açıklama yoksa, desc_idx olmayan kolonlardan birleştir
    if not description:
        skip = {date_idx, amount_idx, balance_idx, receipt_idx, tag_idx}
        parts = [cells[i] for i in range(len(cells)) if i not in skip and cells[i].strip()]
        description = " ".join(parts) if parts else "-"

    # Saat (TEB SAAT kolonu)
    tx_time = None
    time_idx = col_map.get("time")
    if time_idx is not None and time_idx < len(cells):
        raw_time = str(cells[time_idx]).strip()
        if re.match(r"\d{1,2}:\d{2}", raw_time):
            tx_time = raw_time[:5]  # "HH:MM"

    tx_type = "income" if amount >= 0 else "expense"
    tx_hash = compute_tx_hash(tx_date, receipt_no, amount, description)

    return ParsedTransaction(
        date=tx_date,
        receipt_no=receipt_no,
        description=description,
        amount=amount,
        balance=balance,
        type=tx_type,
        tx_hash=tx_hash,
        time=tx_time,
    )


# ─── VakıfBank Ayrıştırıcı ───────────────────────────────

def _parse_vakifbank_text(text: str) -> List[ParsedTransaction]:
    """VakıfBank formatı: TARİH SAAT İŞLEM_NO MİKTAR BAKİYE İŞLEM_ADI + çok satırlı açıklama."""
    transactions = []
    lines = text.split("\n")

    # VakıfBank satır deseni: DD.MM.YYYY HH:MM İŞLEM_NO TUTAR BAKİYE [AÇIKLAMA]
    # Açıklama opsiyonel: bazı işlemlerde açıklama bir sonraki satırda olabilir
    vakif_re = re.compile(
        r"(\d{2}\.\d{2}\.\d{4})\s+"     # Tarih
        r"\d{2}:\d{2}\s+"               # Saat (atla)
        r"(\d{13,20})\s+"               # İşlem No (13-20 haneli)
        r"(-?[\d.,]+)\s+"               # Miktar
        r"(-?[\d.,]+)"                   # Bakiye
        r"(?:\s+(.+))?"                  # İşlem Adı (opsiyonel, satırın geri kalanı)
    )

    i = 0
    while i < len(lines):
        line = lines[i].strip()
        m = vakif_re.match(line)
        if not m:
            i += 1
            continue

        tx_date = parse_date_tr(m.group(1))
        if not tx_date:
            i += 1
            continue

        receipt_no = m.group(2)
        amount = parse_turkish_number(m.group(3))
        balance = parse_turkish_number(m.group(4))
        description = (m.group(5) or "").strip()

        if amount is None:
            i += 1
            continue

        # Çok satırlı açıklama: sonraki satırlar tarihle başlamıyorsa birleştir
        i += 1
        while i < len(lines):
            next_line = lines[i].strip()
            if not next_line:
                i += 1
                continue
            # Yeni işlem satırı veya header/footer mı?
            if re.match(r"\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}", next_line):
                break
            if re.match(r"(Hesap\s+Hareketleri|TAR[İI]H\s+SAAT|Hesap\s+No|Sayfa\s*:)", next_line, re.IGNORECASE):
                break
            description += " " + next_line
            i += 1

        # Açıklamayı temizle
        description = re.sub(r"\s+", " ", description).strip()
        if not description:
            description = "-"

        tx_type = "income" if amount >= 0 else "expense"
        tx_hash = compute_tx_hash(tx_date, receipt_no, amount, description)

        transactions.append(ParsedTransaction(
            date=tx_date,
            receipt_no=receipt_no,
            description=description,
            amount=amount,
            balance=balance,
            type=tx_type,
            tx_hash=tx_hash,
        ))

    return transactions


# ─── Metin Tabanlı Ayrıştırıcı ──────────────────────────

def _parse_text_rows(text: str) -> List[ParsedTransaction]:
    """Metin tabanlı satır satır ayrıştırma (tüm bankalar)."""
    transactions = []
    lines = text.split("\n")

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # DD.MM.YYYY veya DD/MM/YYYY ile başlamalı
        date_match = re.match(r"(\d{2}[./]\d{2}[./]\d{4})\s+", line)
        if not date_match:
            continue

        tx_date = parse_date_tr(date_match.group(1))
        if not tx_date:
            continue

        rest = line[date_match.end():].strip()

        # Footer satırlarını atla
        if re.match(r"(Bor[çc]|Alacak|Toplam|Sayfa)", rest, re.IGNORECASE):
            continue

        # Satır sonundan tutar ve bakiye çıkar
        nums, remaining = _extract_trailing_numbers(rest)
        if not nums:
            continue

        if len(nums) >= 2:
            amount = _smart_parse_number(nums[-2])
            balance = _smart_parse_number(nums[-1])
        else:
            amount = _smart_parse_number(nums[-1])
            balance = None

        if amount is None:
            continue

        # Fiş No / İşlem No — başlangıçtaki alfanümerik token
        receipt_match = re.match(r"([A-Za-z]*\d{3,})\s+", remaining)
        receipt_no = None
        if receipt_match:
            receipt_no = receipt_match.group(1)
            description = remaining[receipt_match.end():].strip()
        else:
            # Saat kalıbını atla (HH:MM:SS veya HH:MM)
            time_match = re.match(r"\d{2}:\d{2}(?::\d{2})?\s+", remaining)
            if time_match:
                remaining = remaining[time_match.end():].strip()
                # Saatten sonra fiş no / işlem no kontrol et
                receipt_match2 = re.match(r"([A-Za-z]*\d{3,})\s+", remaining)
                if receipt_match2:
                    receipt_no = receipt_match2.group(1)
                    remaining = remaining[receipt_match2.end():].strip()
            description = remaining.strip()

        # Sondaki sayısal artıkları temizle
        description = re.sub(r"\s+[-+]?[\d.,]+\s*$", "", description).strip()

        if not description:
            description = "-"

        tx_type = "income" if amount >= 0 else "expense"
        tx_hash = compute_tx_hash(tx_date, receipt_no, amount, description)

        transactions.append(ParsedTransaction(
            date=tx_date,
            receipt_no=receipt_no,
            description=description,
            amount=amount,
            balance=balance,
            type=tx_type,
            tx_hash=tx_hash,
        ))

    return transactions


def _repair_balance_gaps(transactions: List[ParsedTransaction]) -> List[ParsedTransaction]:
    """Bakiye zincirindeki boşluklardan eksik işlemleri kurtarır.

    OCR'lı PDF'lerde çok satırlı açıklamaların tarihi bozulabilir
    ve bu işlemler parse edilemez. Bakiye zinciri tutarsızlığından
    eksik tutarı hesaplayarak kurtarma yaparız.

    Formül: prev.balance + missing_amount + cur.amount = cur.balance
    → missing_amount = cur.balance - cur.amount - prev.balance
    """
    if len(transactions) < 2:
        return transactions

    repaired = list(transactions)
    inserts = []

    for i in range(1, len(repaired)):
        prev = repaired[i - 1]
        cur = repaired[i]

        if prev.balance is None or cur.balance is None:
            continue

        expected_balance = prev.balance + cur.amount
        if abs(expected_balance - cur.balance) < 0.02:
            continue  # Zincir tutarlı

        # Boşluk var — eksik işlemi hesapla
        missing_amount = cur.balance - cur.amount - prev.balance

        if abs(missing_amount) < 0.01:
            continue  # Yuvarlama hatası

        # Eksik işlem tarihi: cur ile aynı (genellikle aynı gün)
        tx_date = cur.date
        missing_balance = prev.balance + missing_amount

        description = "Bakiye zincirinden kurtarılan işlem"
        tx_type = "income" if missing_amount >= 0 else "expense"
        tx_hash = compute_tx_hash(tx_date, None, missing_amount, description)

        inserts.append((i, ParsedTransaction(
            date=tx_date,
            receipt_no=None,
            description=description,
            amount=missing_amount,
            balance=missing_balance,
            type=tx_type,
            tx_hash=tx_hash,
        )))

    # Eklemeleri ters sırada yap (indeksler kaymasın)
    for idx, tx in reversed(inserts):
        repaired.insert(idx, tx)

    return repaired


# ─── Excel Parser ────────────────────────────────────────

def parse_excel(file_path: str) -> ParseResult:
    """Excel (.xlsx) banka ekstresi ayrıştır (Halkbank vb.)."""
    from openpyxl import load_workbook

    result = ParseResult()
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return result

    # ─── Header meta bilgilerini çıkar (Halkbank: satır 0-22) ───
    for i, row in enumerate(rows[:25]):
        if row is None:
            continue
        row_text = " ".join(str(c) for c in row if c)

        # IBAN — BBAN alfanümerik olabilir (Halkbank hesap no'da harf: "2L001751" →
        # IBAN "TR37...002L001751"); yalnız-rakam desen ("TR\d{24}") bunları kaçırıyordu.
        if not result.header.iban:
            iban_clean = re.sub(r"(?<=[A-Z0-9])\s+(?=[A-Z0-9])", "", row_text, flags=re.IGNORECASE)
            m = re.search(r"(TR\d{2}[A-Z0-9]{22})", iban_clean, re.IGNORECASE)
            if m:
                result.header.iban = m.group(1).upper()

        # Döviz
        if not result.header.currency:
            m = re.search(r"(TRY|TL|EUR|USD)\b", row_text, re.IGNORECASE)
            if m:
                cur = m.group(1).upper()
                result.header.currency = "TRY" if cur in ("TL", "TRY") else cur

        # Şube
        if not result.header.branch_name:
            m = re.search(r"[ŞS]ube\s*:?\s*(.+)", row_text)
            if m:
                result.header.branch_name = m.group(1).strip()

    # ─── Veri header satırını bul ───
    header_idx = -1
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c else "" for c in row]
        joined = " ".join(cells)
        if ("tarih" in joined and
            ("tutar" in joined or "borç" in joined or "alacak" in joined or "bakiye" in joined)):
            header_idx = i
            break

    if header_idx < 0:
        return result

    header_cells = [str(c).strip().lower() if c else "" for c in rows[header_idx]]

    # Kolon indekslerini bul
    col_map = {}
    for i, h in enumerate(header_cells):
        if "tarih" in h and "date" not in col_map:
            # İlk tarih kolonu (Halkbank: "İşlem Tarihi", ikincisi "Hesaba Giriş Tarihi")
            col_map["date"] = i
        elif ("açıklama" in h or "aciklama" in h) and "desc" not in col_map:
            col_map["desc"] = i
        elif ("tutar" in h or "miktar" in h) and "amount" not in col_map:
            col_map["amount"] = i
        elif ("bakiye" in h or "yeni bakiye" in h) and "balance" not in col_map:
            col_map["balance"] = i
        elif ("fiş" in h or "fis" in h or "dekont" in h):
            col_map["receipt"] = i
        elif "borç" in h or "borc" in h:
            col_map["debit"] = i
        elif "alacak" in h:
            col_map["credit"] = i

    if "date" not in col_map:
        return result

    # ─── Veri satırlarını ayrıştır ───
    for row in rows[header_idx + 1:]:
        cells = list(row)
        if not cells or all(c is None for c in cells):
            continue

        # Tarih
        raw_date = cells[col_map["date"]] if col_map["date"] < len(cells) else None
        if raw_date is None:
            continue

        if isinstance(raw_date, datetime):
            tx_date = raw_date.date()
        elif isinstance(raw_date, date):
            tx_date = raw_date
        elif isinstance(raw_date, str):
            tx_date = parse_date_tr(raw_date.strip())
            if not tx_date:
                continue
        else:
            continue

        # Fiş No
        receipt_no = None
        if "receipt" in col_map and col_map["receipt"] < len(cells):
            v = cells[col_map["receipt"]]
            receipt_no = str(v).strip() if v else None

        # Açıklama
        description = ""
        if "desc" in col_map and col_map["desc"] < len(cells):
            v = cells[col_map["desc"]]
            description = str(v).strip() if v else ""

        # Tutar
        amount = None
        if "amount" in col_map and col_map["amount"] < len(cells):
            v = cells[col_map["amount"]]
            if isinstance(v, (int, float)):
                amount = float(v)
            elif isinstance(v, str):
                amount = parse_turkish_number(v)
        elif "debit" in col_map and "credit" in col_map:
            debit_val = cells[col_map["debit"]] if col_map["debit"] < len(cells) else None
            credit_val = cells[col_map["credit"]] if col_map["credit"] < len(cells) else None
            if debit_val and isinstance(debit_val, (int, float)) and debit_val != 0:
                amount = -abs(float(debit_val))
            elif credit_val and isinstance(credit_val, (int, float)) and credit_val != 0:
                amount = abs(float(credit_val))
            elif debit_val and isinstance(debit_val, str):
                parsed = parse_turkish_number(debit_val)
                if parsed and parsed != 0:
                    amount = -abs(parsed)
            elif credit_val and isinstance(credit_val, str):
                parsed = parse_turkish_number(credit_val)
                if parsed and parsed != 0:
                    amount = abs(parsed)

        if amount is None:
            continue

        # Bakiye
        balance = None
        if "balance" in col_map and col_map["balance"] < len(cells):
            v = cells[col_map["balance"]]
            if isinstance(v, (int, float)):
                balance = float(v)
            elif isinstance(v, str):
                balance = parse_turkish_number(v)

        tx_type = "income" if amount >= 0 else "expense"
        tx_hash = compute_tx_hash(tx_date, receipt_no, amount, description)

        result.transactions.append(ParsedTransaction(
            date=tx_date,
            receipt_no=receipt_no,
            description=description,
            amount=amount,
            balance=balance,
            type=tx_type,
            tx_hash=tx_hash,
        ))

    # Kronolojik sıraya getir (eskiden yeniye)
    _ensure_chronological_order(result.transactions)

    return result
