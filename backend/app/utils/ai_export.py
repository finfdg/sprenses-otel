"""Asistan tablosunu Excel/PDF'e aktarma yardımcıları.

Frontend, asistan yanıtındaki markdown tablosunu (başlıklar + satırlar) çıkarıp
`/api/ai/disa-aktar`'a gönderir; bu modül openpyxl (xlsx) veya reportlab (pdf) ile
indirilebilir dosya üretir. PDF Türkçe + ₺ için DejaVu font kullanır (pdf_fonts).
"""
import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.utils.pdf_fonts import register_turkish_fonts

_LACIVERT = "1B2B45"   # tema birincil (başlık zemini)
_KREM = "F4F1EA"       # tema alternatif satır


def build_xlsx(baslik: str, basliklar: List[str], satirlar: List[List[str]]) -> bytes:
    """openpyxl ile .xlsx üret → bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"

    ws.append(basliklar)
    header_fill = PatternFill("solid", fgColor=_LACIVERT)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in satirlar:
        ws.append(row)

    # Kolon genişliklerini içeriğe göre ayarla
    for i in range(1, len(basliklar) + 1):
        col = get_column_letter(i)
        lengths = [len(str(basliklar[i - 1]))]
        for row in satirlar:
            if i - 1 < len(row):
                lengths.append(len(str(row[i - 1])))
        ws.column_dimensions[col].width = min(max(lengths) + 2, 60)

    ws.freeze_panes = "A2"  # başlık satırı sabit

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(baslik: str, basliklar: List[str], satirlar: List[List[str]]) -> bytes:
    """reportlab ile .pdf üret (Türkçe font, hücreler sarılır) → bytes."""
    base_font, bold_font = register_turkish_fonts()

    title_style = ParagraphStyle("baslik", fontName=bold_font, fontSize=13, spaceAfter=6)
    head_style = ParagraphStyle("head", fontName=bold_font, fontSize=8, textColor=colors.white)
    cell_style = ParagraphStyle("cell", fontName=base_font, fontSize=8, leading=10)

    # Her hücreyi Paragraph'a sar → uzun metin sarılır, Türkçe/₺ düzgün render olur
    header_cells = [Paragraph(str(h), head_style) for h in basliklar]
    data = [header_cells]
    for row in satirlar:
        cells = [Paragraph(str(row[i]) if i < len(row) else "", cell_style)
                 for i in range(len(basliklar))]
        data.append(cells)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + _LACIVERT)),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#" + _KREM)]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=28, bottomMargin=28, leftMargin=28, rightMargin=28,
    )
    doc.build([Paragraph(baslik or "Rapor", title_style), Spacer(1, 6), table])
    return buf.getvalue()
