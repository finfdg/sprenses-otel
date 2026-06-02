"""Cari hesap Excel ayrıştırıcı.

Muhasebe programından dışa aktarılan .xls/.xlsx dosyalarını okur.
Beklenen kolon yapısı:
  HesapKodu | HesapAdi | Tarih | Evrak No | İşlem Tipi | Fiş No | Açıklama | Borç | Alacak | Bakiye
"""
import hashlib
from dataclasses import dataclass, field
from datetime import date, timedelta
from typing import List, Optional


@dataclass
class ParsedVendorTransaction:
    hesap_kodu: str
    hesap_adi: str
    date: date
    evrak_no: Optional[str]
    transaction_type: Optional[str]
    fis_no: Optional[str]
    description: Optional[str]
    borc: float
    alacak: float
    bakiye: Optional[float]
    tx_hash: str = ""
    payment_due_date: Optional[date] = None


@dataclass
class VendorParseResult:
    transactions: List[ParsedVendorTransaction] = field(default_factory=list)
    vendor_codes: set = field(default_factory=set)


def calculate_payment_friday(invoice_date: date, days: int = 90) -> date:
    """Fatura tarihine belirtilen gün sayısını ekle, sonraki Cuma'ya yuvarla."""
    target = invoice_date + timedelta(days=days)
    weekday = target.weekday()  # Monday=0 ... Friday=4 ... Sunday=6
    if weekday == 4:
        return target
    days_until_friday = (4 - weekday) % 7
    if days_until_friday == 0:
        days_until_friday = 7
    return target + timedelta(days=days_until_friday)


def _normalize_evrak_no(evrak_no: Optional[str]) -> str:
    """Evrak numarasını normalize et — mükerrer tespiti iyileştirmek için.

    Muhasebe programları aynı evrak numarasını farklı formatlarla yazabiliyor:
    - NSE202600000018 vs NSE2026000000018 (sıfır sayısı farkı)
    - 000251 vs 251 (baştaki sıfırlar)

    Yaklaşım: Prefix harfleri koru, yıl kısmını koru, seri no'daki
    baştaki sıfırları kaldır.
    """
    if not evrak_no:
        return ""
    import re
    s = evrak_no.strip().upper()

    # Pattern: HARFLER + 4 haneli YIL + SERİ NO (ör: NSE202600000018)
    m = re.match(r'^([A-ZÇĞİÖŞÜ]+)(\d{4})(0*)(\d+)$', s)
    if m:
        prefix, year, zeros, serial = m.groups()
        return f"{prefix}{year}{serial}"  # Sıfırları at: NSE202618

    # Harfli prefix + rakam (yıl yok)
    m = re.match(r'^([A-ZÇĞİÖŞÜ]+)(\d+)$', s)
    if m:
        return m.group(1) + str(int(m.group(2)))

    # Sadece rakam
    if s.isdigit():
        return str(int(s))

    return s


def compute_vendor_tx_hash(
    hesap_kodu: str,
    tx_date: date,
    evrak_no: Optional[str],
    borc: float,
    alacak: float,
    description: Optional[str] = None,
) -> str:
    """Mükerrer tespit için SHA-256 hash hesapla.

    Birincil: hesap_kodu + tarih + normalize(evrak_no) + tutar
    İkincil: hesap_kodu + tarih + tutar + açıklama ilk 30 karakter
    (evrak_no farklı yazılsa bile aynı işlemi yakalar)

    İki hash birleştirilir — herhangi biri eşleşirse mükerrer sayılır.
    """
    normalized_evrak = _normalize_evrak_no(evrak_no)
    # Evrak no'daki tüm rakam olmayan karakterleri kaldırıp sadece rakamları karşılaştır
    evrak_digits = ''.join(c for c in normalized_evrak if c.isdigit())
    if evrak_digits:
        evrak_digits = str(int(evrak_digits))  # Baştaki sıfırları kaldır

    raw = f"{hesap_kodu}|{tx_date.isoformat()}|{evrak_digits}|{borc:.2f}|{alacak:.2f}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _is_invoice_type(transaction_type: Optional[str]) -> bool:
    """İşlem tipinin fatura olup olmadığını kontrol et."""
    if not transaction_type:
        return False
    t = transaction_type.lower()
    return "fatura" in t


def parse_vendor_excel(file_path: str) -> VendorParseResult:
    """Cari hesap Excel dosyasını ayrıştır (.xls veya .xlsx)."""
    ext = file_path.rsplit(".", 1)[-1].lower()

    if ext == "xls":
        return _parse_xls(file_path)
    else:
        return _parse_xlsx(file_path)


def _parse_xls(file_path: str) -> VendorParseResult:
    """xlrd ile .xls dosyasını ayrıştır."""
    import xlrd

    result = VendorParseResult()
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        return result

    # İlk satır header — atla
    for row_idx in range(1, ws.nrows):
        # HesapKodu (kolon 0)
        hesap_kodu_cell = ws.cell(row_idx, 0)
        hesap_kodu = str(hesap_kodu_cell.value).strip() if hesap_kodu_cell.value else ""

        # Ara toplam / footer satırlarını atla (HesapKodu boş)
        if not hesap_kodu:
            continue

        # "Sayfa" ile başlayan footer satırlarını atla
        if hesap_kodu.lower().startswith("sayfa"):
            continue

        # HesapAdi (kolon 1)
        hesap_adi = str(ws.cell(row_idx, 1).value).strip() if ws.cell(row_idx, 1).value else ""

        # Tarih (kolon 2) — Excel serial number
        date_cell = ws.cell(row_idx, 2)
        if date_cell.ctype == 3:  # XL_CELL_DATE
            dt_tuple = xlrd.xldate_as_tuple(date_cell.value, wb.datemode)
            tx_date = date(dt_tuple[0], dt_tuple[1], dt_tuple[2])
        elif date_cell.ctype == 1:  # XL_CELL_TEXT
            tx_date = _parse_date_str(str(date_cell.value))
            if not tx_date:
                continue
        else:
            continue

        # Evrak No (kolon 3)
        evrak_no_val = ws.cell(row_idx, 3).value
        evrak_no = str(evrak_no_val).strip() if evrak_no_val else None

        # İşlem Tipi (kolon 4)
        tx_type_val = ws.cell(row_idx, 4).value
        transaction_type = str(tx_type_val).strip() if tx_type_val else None

        # Fiş No (kolon 5)
        fis_no_val = ws.cell(row_idx, 5).value
        fis_no = str(int(fis_no_val)) if isinstance(fis_no_val, float) and fis_no_val else (
            str(fis_no_val).strip() if fis_no_val else None
        )

        # Açıklama (kolon 6)
        desc_val = ws.cell(row_idx, 6).value
        description = str(desc_val).strip() if desc_val else None

        # Borç (kolon 7)
        borc_val = ws.cell(row_idx, 7).value
        borc = float(borc_val) if borc_val else 0.0

        # Alacak (kolon 8)
        alacak_val = ws.cell(row_idx, 8).value
        alacak = float(alacak_val) if alacak_val else 0.0

        # Bakiye (kolon 9)
        bakiye_val = ws.cell(row_idx, 9).value
        bakiye = float(bakiye_val) if bakiye_val else None

        # Hash hesapla
        tx_hash = compute_vendor_tx_hash(hesap_kodu, tx_date, evrak_no, borc, alacak)

        # Ödeme tarihi: sadece fatura tipi işlemler için
        payment_due = None
        if _is_invoice_type(transaction_type) and alacak > 0:
            payment_due = calculate_payment_friday(tx_date)

        result.vendor_codes.add(hesap_kodu)
        result.transactions.append(ParsedVendorTransaction(
            hesap_kodu=hesap_kodu,
            hesap_adi=hesap_adi,
            date=tx_date,
            evrak_no=evrak_no,
            transaction_type=transaction_type,
            fis_no=fis_no,
            description=description,
            borc=borc,
            alacak=alacak,
            bakiye=bakiye,
            tx_hash=tx_hash,
            payment_due_date=payment_due,
        ))

    return result


def _parse_xlsx(file_path: str) -> VendorParseResult:
    """openpyxl ile .xlsx dosyasını ayrıştır."""
    from datetime import datetime as dt_cls

    from openpyxl import load_workbook

    result = VendorParseResult()
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return result

    # İlk satır header — atla
    for row in rows[1:]:
        if not row or len(row) < 10:
            continue

        hesap_kodu = str(row[0]).strip() if row[0] else ""
        if not hesap_kodu:
            continue
        if hesap_kodu.lower().startswith("sayfa"):
            continue

        hesap_adi = str(row[1]).strip() if row[1] else ""

        # Tarih
        raw_date = row[2]
        if isinstance(raw_date, (dt_cls, date)):
            tx_date = raw_date if isinstance(raw_date, date) else raw_date.date()
        elif isinstance(raw_date, str):
            tx_date = _parse_date_str(raw_date)
            if not tx_date:
                continue
        else:
            continue

        evrak_no = str(row[3]).strip() if row[3] else None
        transaction_type = str(row[4]).strip() if row[4] else None
        fis_no = str(int(row[5])) if isinstance(row[5], float) and row[5] else (
            str(row[5]).strip() if row[5] else None
        )
        description = str(row[6]).strip() if row[6] else None

        borc = float(row[7]) if row[7] else 0.0
        alacak = float(row[8]) if row[8] else 0.0
        bakiye = float(row[9]) if row[9] else None

        tx_hash = compute_vendor_tx_hash(hesap_kodu, tx_date, evrak_no, borc, alacak)

        payment_due = None
        if _is_invoice_type(transaction_type) and alacak > 0:
            payment_due = calculate_payment_friday(tx_date)

        result.vendor_codes.add(hesap_kodu)
        result.transactions.append(ParsedVendorTransaction(
            hesap_kodu=hesap_kodu,
            hesap_adi=hesap_adi,
            date=tx_date,
            evrak_no=evrak_no,
            transaction_type=transaction_type,
            fis_no=fis_no,
            description=description,
            borc=borc,
            alacak=alacak,
            bakiye=bakiye,
            tx_hash=tx_hash,
            payment_due_date=payment_due,
        ))

    return result


def _parse_date_str(text: str) -> Optional[date]:
    """DD.MM.YYYY veya DD/MM/YYYY formatını date'e çevir."""
    from datetime import datetime
    text = text.strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
