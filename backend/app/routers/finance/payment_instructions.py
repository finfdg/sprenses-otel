"""Ödeme talimat listeleri — cari ödemeleri için toplu talimat hazırlama.

Kullanıcı carileri seçip bir listeye ekler; her kalemin tutarı carinin
bakiyesinden (net borç) otomatik gelir ama manuel düzenlenebilir. Liste kalıcıdır,
Excel/PDF olarak dışa aktarılabilir (export_instruction.py).
"""

import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.middleware.auth import require_permission
from app.middleware.rate_limit import get_client_ip
from app.models.payment_instruction import (
    PaymentInstructionItem,
    PaymentInstructionList,
)
from app.models.user import User
from app.models.vendor import Vendor
from app.schemas.payment_instruction import (
    BulkAddItemsRequest,
    PaymentItemResponse,
    PaymentItemUpdate,
    PaymentListCreate,
    PaymentListResponse,
    PaymentListUpdate,
)
from app.utils.audit import log_action

router = APIRouter(prefix="/payment-instructions")

MAX_ITEMS = 1000


def _valid_vendor_ids(db: Session, vendor_ids) -> set:
    """Verilen vendor_id'lerden DB'de gerçekten var olanları döndür.

    Var olmayan (silinmiş/geçersiz) vendor_id ile kalem eklenirse FK ihlali (500)
    oluşur. Bu helper ile geçersiz id'ler kalemde None'a çevrilir; hesap_kodu/adi
    snapshot alanları kalemi korur (modelin vendor FK'si ondelete=SET NULL'dur).
    """
    ids = {v for v in vendor_ids if v is not None}
    if not ids:
        return set()
    rows = db.query(Vendor.id).filter(Vendor.id.in_(ids)).all()
    return {r[0] for r in rows}


def _fmt_try(v: float) -> str:
    """Türkçe para formatı: 1.234.567,89"""
    s = f"{v:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _norm_iban(s) -> Optional[str]:
    """IBAN normalize: büyük harf, boşluksuz. Boşsa None."""
    if not s:
        return None
    v = "".join(str(s).split()).upper()
    return v or None


def _fmt_iban(iban) -> str:
    """IBAN'ı 4'erli gruplayarak okunur göster (TR12 3456 ...)."""
    v = (iban or "").strip()
    return " ".join(v[i:i + 4] for i in range(0, len(v), 4)) if v else ""


# ─── Yardımcı ────────────────────────────────────────────

def _build_list_response(pl: PaymentInstructionList, with_items: bool = True) -> dict:
    """Liste yanıtı oluştur — item_count ve total_amount hesaplanır."""
    items = list(pl.items)
    total = sum(float(it.amount or 0) for it in items)
    creator_name = None
    if pl.creator:
        creator_name = pl.creator.full_name
    return PaymentListResponse(
        id=pl.id,
        name=pl.name,
        description=pl.description,
        status=pl.status,
        item_count=len(items),
        total_amount=round(total, 2),
        created_by=pl.created_by,
        creator_name=creator_name,
        created_at=pl.created_at,
        updated_at=pl.updated_at,
        items=[
            PaymentItemResponse(
                id=it.id,
                vendor_id=it.vendor_id,
                hesap_kodu=it.hesap_kodu,
                hesap_adi=it.hesap_adi,
                amount=float(it.amount or 0),
                balance_snapshot=float(it.balance_snapshot) if it.balance_snapshot is not None else None,
                notes=it.notes,
                sort_order=it.sort_order,
                bank_name=it.bank_name,
                iban=it.iban,
            )
            for it in items
        ] if with_items else [],
    ).model_dump(mode="json")


def _get_list_or_404(db: Session, list_id: int) -> PaymentInstructionList:
    pl = (
        db.query(PaymentInstructionList)
        .options(selectinload(PaymentInstructionList.items))
        .filter(PaymentInstructionList.id == list_id)
        .first()
    )
    if not pl:
        raise HTTPException(status_code=404, detail="Ödeme talimat listesi bulunamadı")
    return pl


def _next_sort_order(pl: PaymentInstructionList) -> int:
    if not pl.items:
        return 0
    return max((it.sort_order for it in pl.items), default=0) + 1


# ─── Liste CRUD ──────────────────────────────────────────

@router.get("/")
def list_instruction_lists(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("finance.cariler", "view")),
):
    """Tüm ödeme talimat listelerini özet olarak getir (kalemsiz)."""
    lists = (
        db.query(PaymentInstructionList)
        .options(selectinload(PaymentInstructionList.items))
        .order_by(PaymentInstructionList.created_at.desc())
        .all()
    )
    return [_build_list_response(pl, with_items=False) for pl in lists]


@router.get("/{list_id}")
def get_instruction_list(
    list_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("finance.cariler", "view")),
):
    """Tek liste detayı (kalemler dahil)."""
    pl = _get_list_or_404(db, list_id)
    return _build_list_response(pl, with_items=True)


@router.post("/", status_code=status.HTTP_201_CREATED)
def create_instruction_list(
    data: PaymentListCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Yeni ödeme talimat listesi oluştur (opsiyonel başlangıç kalemleriyle)."""
    if len(data.items) > MAX_ITEMS:
        raise HTTPException(status_code=400, detail=f"En fazla {MAX_ITEMS} kalem eklenebilir")

    pl = PaymentInstructionList(
        name=data.name.strip(),
        description=data.description,
        created_by=current_user.id,
    )
    db.add(pl)
    db.flush()

    valid_vids = _valid_vendor_ids(db, [it.vendor_id for it in data.items])
    for i, item in enumerate(data.items):
        db.add(PaymentInstructionItem(
            list_id=pl.id,
            vendor_id=item.vendor_id if item.vendor_id in valid_vids else None,
            hesap_kodu=item.hesap_kodu,
            hesap_adi=item.hesap_adi,
            amount=item.amount,
            balance_snapshot=item.balance_snapshot,
            notes=item.notes,
            sort_order=i,
        ))

    log_action(
        db, current_user.id, "create", "payment_instruction_list",
        entity_id=pl.id,
        details=f"Ödeme talimat listesi oluşturuldu: {pl.name} ({len(data.items)} kalem)",
        ip_address=get_client_ip(request),
    )
    db.commit()
    db.refresh(pl)
    return _build_list_response(pl, with_items=True)


@router.patch("/{list_id}")
def update_instruction_list(
    list_id: int,
    data: PaymentListUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Liste başlığını güncelle (ad, açıklama, durum)."""
    pl = _get_list_or_404(db, list_id)
    update_data = data.model_dump(exclude_unset=True)
    if "name" in update_data and update_data["name"]:
        update_data["name"] = update_data["name"].strip()
    for key, value in update_data.items():
        setattr(pl, key, value)

    log_action(
        db, current_user.id, "update", "payment_instruction_list",
        entity_id=pl.id,
        details=f"Ödeme talimat listesi güncellendi: {pl.name}",
        ip_address=get_client_ip(request),
    )
    db.commit()
    db.refresh(pl)
    return _build_list_response(pl, with_items=True)


@router.delete("/{list_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_instruction_list(
    list_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Listeyi (ve kalemlerini) sil."""
    pl = _get_list_or_404(db, list_id)
    name = pl.name
    db.delete(pl)
    log_action(
        db, current_user.id, "delete", "payment_instruction_list",
        entity_id=list_id,
        details=f"Ödeme talimat listesi silindi: {name}",
        ip_address=get_client_ip(request),
    )
    db.commit()


# ─── Kalem Yönetimi ──────────────────────────────────────

@router.post("/{list_id}/items")
def add_items(
    list_id: int,
    body: BulkAddItemsRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Listeye bir veya birden fazla cari kalemi ekle.

    Aynı vendor_id zaten listedeyse atlanır (mükerrer cari engellenir).
    """
    pl = _get_list_or_404(db, list_id)
    existing_vendor_ids = {it.vendor_id for it in pl.items if it.vendor_id is not None}
    if len(pl.items) + len(body.items) > MAX_ITEMS:
        raise HTTPException(status_code=400, detail=f"En fazla {MAX_ITEMS} kalem eklenebilir")

    so = _next_sort_order(pl)
    valid_vids = _valid_vendor_ids(db, [it.vendor_id for it in body.items])

    # Carilerin varsayılan banka/IBAN'ı — kalemde belirtilmemişse otomatik gelir
    from app.models.vendor_bank_account import VendorBankAccount
    default_bank = {}
    add_vids = [v for v in (it.vendor_id for it in body.items) if v]
    if add_vids:
        for ba in (
            db.query(VendorBankAccount)
            .filter(VendorBankAccount.vendor_id.in_(add_vids))
            .order_by(VendorBankAccount.is_default.desc(), VendorBankAccount.sort_order)
            .all()
        ):
            if ba.vendor_id not in default_bank:  # ilk (varsayılan) kazanır
                default_bank[ba.vendor_id] = (ba.bank_name, ba.iban)

    added = 0
    skipped = 0
    for item in body.items:
        if item.vendor_id is not None and item.vendor_id in existing_vendor_ids:
            skipped += 1
            continue
        b_name, b_iban = item.bank_name, _norm_iban(item.iban)
        if b_name is None and b_iban is None and item.vendor_id in default_bank:
            b_name, b_iban = default_bank[item.vendor_id]
        db.add(PaymentInstructionItem(
            list_id=pl.id,
            vendor_id=item.vendor_id if item.vendor_id in valid_vids else None,
            hesap_kodu=item.hesap_kodu,
            hesap_adi=item.hesap_adi,
            amount=item.amount,
            balance_snapshot=item.balance_snapshot,
            notes=item.notes,
            bank_name=b_name,
            iban=b_iban,
            sort_order=so,
        ))
        if item.vendor_id is not None:
            existing_vendor_ids.add(item.vendor_id)
        so += 1
        added += 1

    if added:
        log_action(
            db, current_user.id, "update", "payment_instruction_list",
            entity_id=pl.id,
            details=f"Talimat listesine {added} kalem eklendi ({skipped} mükerrer atlandı): {pl.name}",
            ip_address=get_client_ip(request),
        )
    db.commit()
    db.refresh(pl)
    result = _build_list_response(pl, with_items=True)
    result["added"] = added
    result["skipped"] = skipped
    return result


@router.patch("/{list_id}/items/{item_id}")
def update_item(
    list_id: int,
    item_id: int,
    data: PaymentItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Kalem tutarını / notunu / sırasını güncelle."""
    item = (
        db.query(PaymentInstructionItem)
        .filter(
            PaymentInstructionItem.id == item_id,
            PaymentInstructionItem.list_id == list_id,
        )
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Kalem bulunamadı")

    update_data = data.model_dump(exclude_unset=True)
    if "iban" in update_data:
        update_data["iban"] = _norm_iban(update_data["iban"])
    for key, value in update_data.items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return PaymentItemResponse(
        id=item.id,
        vendor_id=item.vendor_id,
        hesap_kodu=item.hesap_kodu,
        hesap_adi=item.hesap_adi,
        amount=float(item.amount or 0),
        balance_snapshot=float(item.balance_snapshot) if item.balance_snapshot is not None else None,
        notes=item.notes,
        sort_order=item.sort_order,
        bank_name=item.bank_name,
        iban=item.iban,
    ).model_dump()


@router.delete("/{list_id}/items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_item(
    list_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Kalemi listeden çıkar."""
    item = (
        db.query(PaymentInstructionItem)
        .filter(
            PaymentInstructionItem.id == item_id,
            PaymentInstructionItem.list_id == list_id,
        )
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Kalem bulunamadı")
    db.delete(item)
    db.commit()


# ─── Dışa Aktarma (Excel / PDF) ──────────────────────────

@router.get("/{list_id}/export/excel")
def export_excel(
    list_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("finance.cariler", "view")),
):
    """Ödeme talimat listesini Excel olarak indir."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    pl = _get_list_or_404(db, list_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ödeme Talimatı"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14)

    ws.cell(row=1, column=1, value=pl.name).font = title_font
    ws.cell(row=2, column=1, value=f"Oluşturulma: {pl.created_at.strftime('%d.%m.%Y') if pl.created_at else '-'}")

    headers = ["Sıra", "Hesap Kodu", "Cari Adı", "Banka", "IBAN", "Açıklama", "Ödeme Tutarı (₺)"]
    head_row = 4
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=head_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total = 0.0
    r = head_row + 1
    for i, it in enumerate(pl.items, start=1):
        amt = float(it.amount or 0)
        total += amt
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=it.hesap_kodu or "")
        ws.cell(row=r, column=3, value=it.hesap_adi)
        ws.cell(row=r, column=4, value=it.bank_name or "")
        ws.cell(row=r, column=5, value=_fmt_iban(it.iban))
        ws.cell(row=r, column=6, value=it.notes or "")
        c = ws.cell(row=r, column=7, value=amt)
        c.number_format = "#,##0.00"
        r += 1

    ws.cell(row=r, column=6, value="TOPLAM").font = Font(bold=True)
    tc = ws.cell(row=r, column=7, value=round(total, 2))
    tc.font = Font(bold=True)
    tc.number_format = "#,##0.00"

    widths = [8, 18, 38, 22, 34, 26, 16]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=head_row, column=col).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = "".join(c for c in pl.name if c.isalnum() or c in " -_")[:40].strip() or "talimat"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=odeme-talimati-{list_id}.xlsx"},
    )


@router.get("/{list_id}/export/pdf")
def export_pdf(
    list_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("finance.cariler", "view")),
):
    """Ödeme talimat listesini PDF olarak indir."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    from app.utils.pdf_fonts import register_turkish_fonts

    pl = _get_list_or_404(db, list_id)

    # Türkçe + para birimi sembolü (₺) destekli font — DejaVuSans
    base_font, bold_font = register_turkish_fonts()

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        topMargin=18 * mm, bottomMargin=18 * mm,
        leftMargin=15 * mm, rightMargin=15 * mm,
    )

    title_style = ParagraphStyle("t", fontName=bold_font, fontSize=14, spaceAfter=4)
    sub_style = ParagraphStyle("s", fontName=base_font, fontSize=9, textColor=colors.grey, spaceAfter=10)

    elems = [
        Paragraph("Ödeme Talimat Listesi", title_style),
        Paragraph(
            f"{pl.name} &nbsp;·&nbsp; Oluşturulma: "
            f"{pl.created_at.strftime('%d.%m.%Y') if pl.created_at else '-'}",
            sub_style,
        ),
        Spacer(1, 4),
    ]

    data = [["#", "Hesap Kodu", "Cari Adı", "Banka", "IBAN", "Tutar (₺)"]]
    total = 0.0
    for i, it in enumerate(pl.items, start=1):
        amt = float(it.amount or 0)
        total += amt
        data.append([
            str(i), it.hesap_kodu or "", it.hesap_adi,
            it.bank_name or "", _fmt_iban(it.iban), _fmt_try(amt),
        ])
    data.append(["", "", "", "", "TOPLAM", _fmt_try(round(total, 2))])

    table = Table(data, colWidths=[8 * mm, 24 * mm, 42 * mm, 22 * mm, 52 * mm, 30 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTNAME", (0, 1), (-1, -1), base_font),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D9488")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (5, 0), (5, -1), "RIGHT"),
        ("ALIGN", (4, -1), (4, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F3F4F6")]),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#0D9488")),
        ("FONTNAME", (0, -1), (-1, -1), bold_font),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#374151")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elems.append(table)

    doc.build(elems)
    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=odeme-talimati-{list_id}.pdf"},
    )
