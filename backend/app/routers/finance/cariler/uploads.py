"""Cari hesap dosya yükleme ve yükleme geçmişi."""

import asyncio
import os
import uuid

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Request, UploadFile, status
from sqlalchemy import desc, or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.middleware.auth import require_permission
from app.middleware.rate_limit import get_client_ip, upload_limiter
from app.models.finance_event import FinanceEvent
from app.models.user import User
from app.models.vendor import Vendor
from app.models.vendor_transaction import VendorTransaction
from app.models.vendor_upload import VendorUpload
from app.schemas.vendor import (
    BulkDeleteRequest,
    BulkDeleteResult,
    RemovalCandidate,
    VendorUploadResponse,
    VendorUploadResult,
)
from app.utils.audit import log_action
from app.utils.file_validation import validate_upload_file
from app.constants import BroadcastModule
from app.utils.finance_broadcast import broadcast_finance_update
from app.utils.finance_event_service import finance_event_svc
from app.utils.sync_vendor_fifo import sync_vendor_finance_events
from app.utils.vendor_parser import calculate_payment_friday, parse_vendor_excel

from ._helpers import UPLOAD_DIR, _ensure_upload_dir, logger

# Korunan dept_status değerleri — manuel iş yapılmış kayıtlar otomatik silme adayı olamaz
_PROTECTED_DEPT_STATUSES = ("assigned", "approved")

router = APIRouter()


def _compute_removal_candidates(
    db: Session,
    vendor_map: dict,
    parsed_transactions: list,
) -> list:
    """Excel kapsamında olduğu halde Excel'de bulunmayan kayıtları belirle.

    Kapsam: (Excel'deki vendor_id'ler) ∩ (Excel'in min/max tarih aralığı)
    Korunan kayıtlar (otomatik atlanır):
      - match_number IS NOT NULL          → banka/çek ile eşleşmiş
      - dept_status IN ('assigned','approved') → departmana atanmış/onaylanmış
      - finance_events.is_matched = TRUE   → karşı tarafla bağlanmış
    """
    if not vendor_map or not parsed_transactions:
        return []

    vendor_ids = [v.id for v in vendor_map.values()]

    # Excel'in tarih aralığı
    excel_dates = [tx.date for tx in parsed_transactions if tx.date]
    if not excel_dates:
        return []
    min_date = min(excel_dates)
    max_date = max(excel_dates)

    # Excel'deki tüm (vendor_id, tx_hash) ikilileri — bu kümede olmayanlar adaydır
    excel_keys = set()
    for tx in parsed_transactions:
        v = vendor_map.get(tx.hesap_kodu)
        if v:
            excel_keys.add((v.id, tx.tx_hash))

    # Kapsamdaki DB kayıtlarını çek (korunanları SQL seviyesinde dışla)
    # Not: dept_status NULL ise NOT IN UNKNOWN döner → OR ile NULL'ı da kabul et
    rows = (
        db.query(VendorTransaction, Vendor)
        .join(Vendor, Vendor.id == VendorTransaction.vendor_id)
        .filter(
            VendorTransaction.vendor_id.in_(vendor_ids),
            VendorTransaction.date >= min_date,
            VendorTransaction.date <= max_date,
            VendorTransaction.match_number.is_(None),
            or_(
                VendorTransaction.dept_status.is_(None),
                ~VendorTransaction.dept_status.in_(_PROTECTED_DEPT_STATUSES),
            ),
        )
        .all()
    )

    if not rows:
        return []

    # finance_events üzerinde eşleşmiş olanları (is_matched=true) ayrıca filtrele
    candidate_ids = [vtx.id for vtx, _ in rows]
    matched_event_ids = {
        row[0]
        for row in db.query(FinanceEvent.source_id)
        .filter(
            FinanceEvent.source_type == "vendor_payment",
            FinanceEvent.source_id.in_(candidate_ids),
            FinanceEvent.is_matched.is_(True),
        )
        .all()
    }

    candidates = []
    for vtx, vendor in rows:
        if vtx.id in matched_event_ids:
            continue
        if (vtx.vendor_id, vtx.tx_hash) in excel_keys:
            continue
        candidates.append(RemovalCandidate(
            id=vtx.id,
            vendor_id=vtx.vendor_id,
            hesap_kodu=vendor.hesap_kodu,
            hesap_adi=vendor.hesap_adi,
            date=vtx.date,
            evrak_no=vtx.evrak_no,
            transaction_type=vtx.transaction_type,
            description=vtx.description,
            borc=float(vtx.borc),
            alacak=float(vtx.alacak),
            bakiye=float(vtx.bakiye) if vtx.bakiye is not None else None,
        ))

    candidates.sort(key=lambda c: (c.hesap_adi, c.date, c.id))
    return candidates


# ─── Dosya Yükleme ──────────────────────────────────────

@router.post("/upload")
async def upload_vendor_excel(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Cari hesap Excel dosyasını yükle ve ayrıştır."""
    upload_limiter.check(f"upload-{get_client_ip(request)}")
    logger.info("Cari dosya yükleniyor: %s (size=%s)", file.filename, file.size)
    content = await validate_upload_file(file, allowed_types=["excel"])

    ext = os.path.splitext(file.filename or "")[1].lower() or ".xlsx"
    _ensure_upload_dir()
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_name)

    with open(file_path, "wb") as f:
        f.write(content)

    try:
        # CPU-yoğun Excel parse'ı threadpool'a al → event loop bloke olmaz (eşzamanlı istekler beklemez)
        parsed = await asyncio.to_thread(parse_vendor_excel, file_path)
    except Exception as e:
        logger.error("Cari dosya ayrıştırma hatası (%s): %s", file.filename, e, exc_info=True)
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=400, detail="Dosya ayrıştırılamadı. Lütfen geçerli bir cari hesap dosyası yükleyin.")

    if not parsed.transactions:
        # Dosya yapısını analiz et
        header_info = ""
        try:
            if ext == "xls":
                import xlrd
                wb_check = xlrd.open_workbook(file_path)
                ws_check = wb_check.sheet_by_index(0)
                if ws_check.nrows > 0:
                    headers = [str(ws_check.cell(0, c).value).strip() for c in range(min(ws_check.ncols, 12))]
                    header_info = f" Dosyadaki kolonlar: {', '.join(h for h in headers if h)}"
                    logger.debug("Boş dosya: rows=%d, cols=%d, headers=%s", ws_check.nrows, ws_check.ncols, headers)
            else:
                from openpyxl import load_workbook as lb
                wb_check = lb(file_path, read_only=True)
                ws_check = wb_check.active
                first_row = next(ws_check.iter_rows(max_row=1, values_only=True), None)
                if first_row:
                    headers = [str(c).strip() if c else "" for c in first_row[:12]]
                    header_info = f" Dosyadaki kolonlar: {', '.join(h for h in headers if h)}"
                    logger.debug("Boş dosya: headers=%s", headers)
                wb_check.close()
        except Exception as e:
            logger.debug("Dosya başlık analizi başarısız: %s", e)

        logger.warning("İşlem bulunamadı: %s, vendor_codes=%s", file.filename, parsed.vendor_codes)
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(
            status_code=400,
            detail=f"Dosyada işlem bulunamadı. Beklenen kolon sırası: HesapKodu, HesapAdi, Tarih, EvrakNo, İşlemTipi, FişNo, Açıklama, Borç, Alacak, Bakiye.{header_info}",
        )

    # Upload kaydı oluştur
    try:
        upload = VendorUpload(
            file_name=file.filename or unique_name,
            file_url=file_path,
            uploaded_by=current_user.id,
        )
        db.add(upload)
        db.flush()

        # Cari kayıtlarını oluştur/güncelle
        vendor_map = {}
        for code in parsed.vendor_codes:
            vendor = db.query(Vendor).filter(Vendor.hesap_kodu == code).first()
            if not vendor:
                name = ""
                for tx in parsed.transactions:
                    if tx.hesap_kodu == code:
                        name = tx.hesap_adi
                        break
                vendor = Vendor(hesap_kodu=code, hesap_adi=name)
                db.add(vendor)
                db.flush()
            vendor_map[code] = vendor

        # Mevcut hash'leri çek (vendor bazlı)
        existing_hashes = set()
        vendor_ids = [v.id for v in vendor_map.values()]
        if vendor_ids:
            rows = (
                db.query(VendorTransaction.vendor_id, VendorTransaction.tx_hash)
                .filter(VendorTransaction.vendor_id.in_(vendor_ids))
                .all()
            )
            existing_hashes = {(r[0], r[1]) for r in rows}

        new_count = 0
        skipped_count = 0

        for tx in parsed.transactions:
            vendor = vendor_map[tx.hesap_kodu]
            key = (vendor.id, tx.tx_hash)

            if key in existing_hashes:
                skipped_count += 1
                continue

            payment_due = None
            if tx.alacak > 0 and tx.date:
                payment_due = calculate_payment_friday(tx.date, vendor.payment_days)

            vtx = VendorTransaction(
                vendor_id=vendor.id,
                upload_id=upload.id,
                date=tx.date,
                evrak_no=tx.evrak_no,
                transaction_type=tx.transaction_type,
                fis_no=tx.fis_no,
                description=tx.description,
                borc=tx.borc,
                alacak=tx.alacak,
                bakiye=tx.bakiye,
                tx_hash=tx.tx_hash,
                payment_due_date=payment_due,
            )
            db.add(vtx)
            db.flush()
            if payment_due:
                finance_event_svc.upsert_vendor_tx(db, vtx, vendor, float(tx.alacak))
            existing_hashes.add(key)
            new_count += 1

        upload.total_vendors = len(parsed.vendor_codes)
        upload.total_transactions = len(parsed.transactions)
        upload.new_transactions = new_count
        upload.skipped_transactions = skipped_count

        # Excel'de bulunmayan kayıtları belirle (silme adayları)
        removal_candidates = _compute_removal_candidates(db, vendor_map, parsed.transactions)

        details_msg = f"Cari dosya yüklendi: {file.filename} ({new_count} yeni, {skipped_count} mükerrer"
        if removal_candidates:
            details_msg += f", {len(removal_candidates)} silme adayı"
        details_msg += ")"

        log_action(
            db, current_user.id, "create", "vendor_upload",
            entity_id=upload.id,
            details=details_msg,
            ip_address=get_client_ip(request),
        )
        db.commit()
        sync_vendor_finance_events(db)
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("Cari dosya yükleme veritabanı hatası (%s): %s", file.filename, e, exc_info=True)
        raise HTTPException(status_code=500, detail="Dosya yüklenirken bir veritabanı hatası oluştu. Lütfen tekrar deneyin.")

    broadcast_finance_update(background_tasks, BroadcastModule.CARILER, "upload")

    return VendorUploadResult(
        upload_id=upload.id,
        file_name=upload.file_name,
        total_vendors=upload.total_vendors,
        total_transactions=upload.total_transactions,
        new_transactions=new_count,
        skipped_transactions=skipped_count,
        removal_candidates=removal_candidates,
    ).model_dump()


# ─── Yükleme Geçmişi ────────────────────────────────────

@router.get("/uploads")
def list_uploads(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("finance.cariler", "view")),
):
    """Yükleme geçmişini listele."""
    uploads = (
        db.query(VendorUpload, User.first_name, User.last_name)
        .outerjoin(User, VendorUpload.uploaded_by == User.id)
        .order_by(desc(VendorUpload.uploaded_at))
        .all()
    )

    result = []
    for upload, first_name, last_name in uploads:
        uploader_name = None
        if first_name or last_name:
            uploader_name = f"{first_name or ''} {last_name or ''}".strip()

        result.append(VendorUploadResponse(
            id=upload.id,
            file_name=upload.file_name,
            total_vendors=upload.total_vendors,
            total_transactions=upload.total_transactions,
            new_transactions=upload.new_transactions,
            skipped_transactions=upload.skipped_transactions,
            uploaded_by=upload.uploaded_by,
            uploader_name=uploader_name,
            uploaded_at=upload.uploaded_at,
        ).model_dump())

    return result


@router.delete("/uploads/{upload_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_upload(
    upload_id: int,
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Yüklemeyi ve ilişkili işlemleri sil."""
    upload = db.query(VendorUpload).filter(VendorUpload.id == upload_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Yükleme bulunamadı")

    if upload.file_url and os.path.exists(upload.file_url):
        os.remove(upload.file_url)

    try:
        vtx_ids = [
            row[0] for row in
            db.query(VendorTransaction.id).filter(VendorTransaction.upload_id == upload_id).all()
        ]
        for vtx_id in vtx_ids:
            finance_event_svc.invalidate(db, "vendor_payment", vtx_id)

        log_action(
            db, current_user.id, "delete", "vendor_upload",
            entity_id=upload_id,
            details=f"Cari dosya silindi: {upload.file_name} ({len(vtx_ids)} işlem)",
            ip_address=get_client_ip(request),
        )

        db.delete(upload)
        db.commit()

        orphaned = (
            db.query(Vendor)
            .outerjoin(VendorTransaction, Vendor.id == VendorTransaction.vendor_id)
            .filter(VendorTransaction.id.is_(None))
            .all()
        )
        for v in orphaned:
            db.delete(v)
        if orphaned:
            db.commit()
        sync_vendor_finance_events(db)
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error("Cari yükleme silme hatası (upload_id=%s): %s", upload_id, e, exc_info=True)
        raise HTTPException(status_code=500, detail="Yükleme silinirken bir veritabanı hatası oluştu.")

    broadcast_finance_update(background_tasks, BroadcastModule.CARILER, "delete")


# ─── Toplu İşlem Silme (kaynakta olmayan kayıtlar) ──────

@router.post("/transactions/bulk-delete")
def bulk_delete_transactions(
    body: BulkDeleteRequest,
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "use")),
):
    """Excel'de bulunmayan cari işlemleri toplu sil.

    Korumalı kayıtlar (match_number, dept_status assigned/approved, finance_event matched)
    sessizce atlanır — yetkisiz silinmez. İlk yükleme response'undaki adaylar bu kontrollerden
    geçmiş olsa da arada başka kullanıcı manuel iş yapmış olabilir → tekrar kontrol edilir.
    """
    if not body.ids:
        return BulkDeleteResult(deleted=0, skipped=0).model_dump()

    if len(body.ids) > 5000:
        raise HTTPException(status_code=400, detail="Tek seferde en fazla 5000 kayıt silinebilir.")

    deleted = 0
    skipped = 0
    skipped_reasons: list = []

    try:
        rows = db.query(VendorTransaction).filter(VendorTransaction.id.in_(body.ids)).all()
        found_ids = {r.id for r in rows}
        missing_ids = set(body.ids) - found_ids
        if missing_ids:
            skipped += len(missing_ids)
            skipped_reasons.append(f"{len(missing_ids)} kayıt bulunamadı")

        # finance_events üzerinde matched olanları tek sorguda bul
        matched_event_ids = {
            row[0]
            for row in db.query(FinanceEvent.source_id)
            .filter(
                FinanceEvent.source_type == "vendor_payment",
                FinanceEvent.source_id.in_(found_ids),
                FinanceEvent.is_matched.is_(True),
            )
            .all()
        }

        protected_match = 0
        protected_dept = 0
        protected_event = 0

        for vtx in rows:
            if vtx.match_number is not None:
                protected_match += 1
                continue
            if vtx.dept_status in _PROTECTED_DEPT_STATUSES:
                protected_dept += 1
                continue
            if vtx.id in matched_event_ids:
                protected_event += 1
                continue

            finance_event_svc.invalidate(db, "vendor_payment", vtx.id)
            db.delete(vtx)
            deleted += 1

        skipped += protected_match + protected_dept + protected_event
        if protected_match:
            skipped_reasons.append(f"{protected_match} kayıt banka/çek ile eşleşmiş")
        if protected_dept:
            skipped_reasons.append(f"{protected_dept} kayıt departmana atanmış/onaylanmış")
        if protected_event:
            skipped_reasons.append(f"{protected_event} kayıt nakit akımda eşleşmiş")

        if deleted:
            log_action(
                db, current_user.id, "delete", "vendor_transaction",
                entity_id=None,
                details=f"Kaynakta olmayan {deleted} cari işlem silindi (toplu silme). Atlanan: {skipped}",
                ip_address=get_client_ip(request),
            )

        # Yetim cariler temizlenir
        orphaned = (
            db.query(Vendor)
            .outerjoin(VendorTransaction, Vendor.id == VendorTransaction.vendor_id)
            .filter(VendorTransaction.id.is_(None))
            .all()
        )
        for v in orphaned:
            db.delete(v)

        db.commit()
        sync_vendor_finance_events(db)
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("Toplu cari işlem silme hatası: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="İşlemler silinirken bir veritabanı hatası oluştu.")

    if deleted:
        broadcast_finance_update(background_tasks, BroadcastModule.CARILER, "delete")

    return BulkDeleteResult(
        deleted=deleted,
        skipped=skipped,
        skipped_reasons=skipped_reasons,
    ).model_dump()
