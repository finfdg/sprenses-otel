"""Haftalık ödeme planı ve Excel export endpoint'leri."""

from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session

from app.database import get_db
from app.middleware.auth import require_permission
from app.models.user import User
from app.models.vendor import Vendor
from app.models.vendor_transaction import VendorTransaction
from app.services.payment_schedule_service import (  # noqa: F401 — geriye uyumluluk: testler/parmak izi bu yoldan import eder (2026-09-02 çıkarımı)
    _next_friday,
    compute_payment_schedule,
)
from app.services.sync_vendor_fifo import (
    sync_vendor_finance_events,  # MODÜL niteliği: audit_finance_invariants `ps.sync_vendor_finance_events`'i monkeypatch'ler — kaldırma
)

# Hesaplama gövdesi 2026-09-02'de BİREBİR `app/services/payment_schedule_service.py`'ye taşındı
# (katman yönü: router → service → model). Bu dosya yalnız router + ince endpoint + Excel export
# + yeniden dışa verim tutar.

router = APIRouter()


# ─── Haftalık Ödeme Planı ────────────────────────────────

@router.get("/payment-schedule")
def get_payment_schedule(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("finance.cariler", "view")),
):
    """Haftalık ödeme planını getir — net borç bazlı, FIFO kırpmalı.

    Vadesi geçmiş faturalar otomatik olarak sonraki Cuma'ya kaydırılır.
    Her çağrıda finance_events tablosu da senkronize edilir.
    """
    # `sync_vendor_finance_events` modül niteliği çağrı anında okunur → parmak izi ölçümünün
    # monkeypatch'i (API-003 yazma etkisizleştirme) korunur.
    return compute_payment_schedule(db, from_date, to_date, sync_fn=sync_vendor_finance_events)


# ─── Excel Export ────────────────────────────────────────


@router.get("/export/vendors")
def export_vendors_excel(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("finance.cariler", "view")),
):
    """Cari listesini Excel olarak indir."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    vendors = (
        db.query(
            Vendor.hesap_kodu,
            Vendor.hesap_adi,
            Vendor.payment_days,
            sa_func.coalesce(sa_func.sum(VendorTransaction.borc), 0).label("total_borc"),
            sa_func.coalesce(sa_func.sum(VendorTransaction.alacak), 0).label("total_alacak"),
            sa_func.count(VendorTransaction.id).label("tx_count"),
        )
        .outerjoin(VendorTransaction, VendorTransaction.vendor_id == Vendor.id)
        .group_by(Vendor.id, Vendor.hesap_kodu, Vendor.hesap_adi, Vendor.payment_days)
        .order_by(Vendor.hesap_kodu)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cariler"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    number_fmt = '#,##0.00'
    red_font = Font(name="Calibri", color="DC2626")
    green_font = Font(name="Calibri", color="059669")

    headers = ["Hesap Kodu", "Hesap Adı", "Vade (Gün)", "Toplam Borç", "Toplam Alacak", "Bakiye", "İşlem Sayısı"]
    col_widths = [18, 45, 12, 18, 18, 18, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col)].width = width

    for row_idx, v in enumerate(vendors, 2):
        bakiye = float(v.total_borc) - float(v.total_alacak)
        row_data = [
            v.hesap_kodu, v.hesap_adi, v.payment_days,
            float(v.total_borc), float(v.total_alacak), bakiye, v.tx_count,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col in (4, 5, 6):
                cell.number_format = number_fmt
            if col == 6 and bakiye < 0:
                cell.font = red_font
            elif col == 6 and bakiye > 0:
                cell.font = green_font

    total_row = len(vendors) + 2
    ws.cell(row=total_row, column=1, value="TOPLAM").font = Font(bold=True)
    for col in (4, 5, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({chr(64+col)}2:{chr(64+col)}{total_row-1})"
        cell.number_format = number_fmt
        cell.font = Font(bold=True)
        cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cariler.xlsx"},
    )


@router.get("/export/payment-schedule")
def export_payment_schedule_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("finance.cariler", "view")),
):
    """Ödeme planını Excel olarak indir (net borç bazlı, FIFO kırpmalı)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    schedule_data = get_payment_schedule(db=db, _=current_user)

    flat_rows = []
    for group in schedule_data:
        for item in group["items"]:
            flat_rows.append(item)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ödeme Planı"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    number_fmt = '#,##0.00'
    date_fmt = 'DD.MM.YYYY'

    headers = ["Vade Tarihi", "Hesap Kodu", "Hesap Adı", "Evrak No", "İşlem Tipi", "Fatura Tarihi", "Tutar"]
    col_widths = [14, 18, 45, 16, 14, 14, 18]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col)].width = width

    for row_idx, r in enumerate(flat_rows, 2):
        due_date = r.get("payment_due_date", "")
        inv_date = r.get("invoice_date", "")
        try:
            due_date = datetime.strptime(str(due_date), "%Y-%m-%d").date() if due_date else ""
        except (ValueError, TypeError):
            pass
        try:
            inv_date = datetime.strptime(str(inv_date), "%Y-%m-%d").date() if inv_date else ""
        except (ValueError, TypeError):
            pass

        row_data = [
            due_date, r.get("hesap_kodu", ""), r.get("hesap_adi", ""),
            r.get("evrak_no", "") or "", r.get("transaction_type", "") or "",
            inv_date, r.get("amount", 0),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col in (1, 6):
                cell.number_format = date_fmt
            if col == 7:
                cell.number_format = number_fmt

    total_row = len(flat_rows) + 2
    ws.cell(row=total_row, column=1, value="TOPLAM").font = Font(bold=True)
    cell = ws.cell(row=total_row, column=7)
    cell.value = f"=SUM(G2:G{total_row-1})"
    cell.number_format = number_fmt
    cell.font = Font(bold=True)
    cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=odeme-plani.xlsx"},
    )
