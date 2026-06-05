"""Personel Devam Takip (PDKS) — kiosk dönen QR, telefon basışı, yönetici paneli.

Akış:
- Girişteki ekran `/devam/ekran?key=KIOSK_KEY` → `GET /attendance/kiosk/qr` ile
  her ~10sn'de dönen QR gösterir. QR, `PUBLIC/devam?k=<token>` URL'ini taşır.
- Personel telefonun YERLEŞİK kamerasıyla QR'ı okutur → URL açılır → kimlik çerezi
  (pdks_token) + k token doğrulanır → giriş/çıkış kaydedilir.
- Personel kimliği: kişisel `access_token` (kurulum linki bir kez açılınca çerez olur).

Güvenlik:
- Zaman-damgalı token HMAC(SECRET, unix_ts) — geçerlilik = panel `refresh_sec` + GRACE(3sn).
  Bayat ekran görüntüsü işe yaramaz. Evden basma: kiosk QR endpoint'i KIOSK_KEY ister
  (admin-only) → güncel token uzaktan çekilemez. (Canlı video aktarımı sınırı: docs/modules/devam-takip.md)
- Tek kullanım yerine personel-bazlı debounce (çift basışı engeller).
- Yönetici işlemleri require_permission(hr.attendance); kiosk/setup/punch public.
- Bu modül onay akışından muaftır (Sunucu/Yedekleme gibi ops modülü).
"""
import hashlib
import hmac
import io
import json
import math
import os
import secrets
import time
from datetime import date, datetime, timedelta
from typing import List, Optional

import pytz
import segno
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile
from pydantic import BaseModel
from sqlalchemy import desc, func
from sqlalchemy.orm import Session

from app.config import settings
from app.constants import WSEvent
from app.database import get_db
from app.middleware.auth import require_permission
from app.middleware.rate_limit import RateLimiter, get_client_ip
from app.models.approval import (
    ACTION_CANCEL,
    STATUS_PENDING,
    STATUS_RETURNED,
    ApprovalRequest,
)
from app.models.attendance_setting import (
    DEFAULT_REFRESH_SEC,
    MAX_REFRESH_SEC,
    MIN_REFRESH_SEC,
    TOKEN_GRACE_SEC,
    AttendanceSetting,
)
from app.models.audit_log import AuditLog
from app.models.personnel import (
    SOURCE_MANUAL,
    SOURCE_PHONE,
    TYPE_IN,
    TYPE_OUT,
    AttendanceLog,
    Personnel,
)
from app.models.user import User
from app.utils.approval_check import check_approval
from app.utils.approval_service import get_pending_approver_ids, process_action
from app.utils.audit import log_action
from app.utils.file_validation import validate_upload_file
from app.websocket.manager import manager

TZ = pytz.timezone("Europe/Istanbul")
SECRET = settings.secret_key.encode()
PUBLIC_BASE = settings.cors_origins.split(",")[0].strip().rstrip("/")
# Kiosk ekranını yetkilendiren stabil, admin-only anahtar (SECRET'ten türetilir)
KIOSK_KEY = hmac.new(SECRET, b"pdks-kiosk-key", hashlib.sha256).hexdigest()[:24]

PUNCH_DEBOUNCE_SEC = 30    # aynı personel bu sürede tekrar basamaz
COOKIE_NAME = "pdks_token"

punch_limiter = RateLimiter(max_requests=40, window_seconds=60)

router = APIRouter()


# ─── Zaman-damgalı token yardımcıları ────────────────────
# Token = "<unix_ts>.<HMAC(SECRET, ts)>" — üretildiği andan itibaren (refresh + GRACE)
# saniye geçerli. Pencere hizalama yok → "geçerlilik süresi" net ve öngörülebilir
# (ekran görüntüsü/bayat QR bu süre dolunca işe yaramaz).

def _sign_ts(ts: int) -> str:
    return hmac.new(SECRET, f"pdks:{ts}".encode(), hashlib.sha256).hexdigest()[:16]


def _make_token() -> str:
    ts = int(time.time())
    return f"{ts}.{_sign_ts(ts)}"


def _valid_token(token: str, ttl_sec: int) -> bool:
    try:
        ts_str, sig = token.split(".", 1)
        ts = int(ts_str)
    except (ValueError, AttributeError):
        return False
    now = int(time.time())
    # +1sn: saat kayması toleransı (ileri); now-ts>TTL: süresi dolmuş
    if ts > now + 1 or now - ts > ttl_sec:
        return False
    return hmac.compare_digest(sig, _sign_ts(ts))


def _get_refresh(db: Session) -> int:
    """Panelden ayarlanan QR yenileme süresi (DB tek satır); yoksa varsayılan."""
    row = db.query(AttendanceSetting).filter(AttendanceSetting.id == 1).first()
    return row.refresh_sec if row else DEFAULT_REFRESH_SEC


def _ttl_for(refresh: int) -> int:
    """Token geçerliliği = yenileme süresi + grace (taze QR'ı tararken pay)."""
    return refresh + TOKEN_GRACE_SEC


def _set_cookie(response: Response, token: str) -> None:
    is_secure = "https" in settings.cors_origins
    response.set_cookie(
        COOKIE_NAME, token,
        max_age=60 * 60 * 24 * 365, httponly=True, secure=is_secure, samesite="lax", path="/",
    )


def _personnel_from_request(request: Request, db: Session) -> Optional[Personnel]:
    """Personel kimliği: önce X-Pdks-Token başlığı (localStorage'dan), sonra çerez.

    iOS Safari, kameranın açtığı sayfada fetch ile set edilen HttpOnly çerezi her
    zaman taşımıyor; bu yüzden frontend kimliği localStorage'da da tutup başlıkla
    gönderir. Aynı-origin istek olduğundan özel başlık CORS'a takılmaz.
    """
    tok = request.headers.get("X-Pdks-Token") or request.cookies.get(COOKIE_NAME)
    if not tok:
        return None
    return db.query(Personnel).filter(
        Personnel.access_token == tok, Personnel.is_active.is_(True)
    ).first()


def _last_log(db: Session, personnel_id: int) -> Optional[AttendanceLog]:
    return (
        db.query(AttendanceLog)
        .filter(AttendanceLog.personnel_id == personnel_id, AttendanceLog.deleted_at.is_(None))
        .order_by(desc(AttendanceLog.punched_at))
        .first()
    )


def _localize(dt: datetime) -> datetime:
    """Naive datetime'ı Europe/Istanbul'a yerelleştir (tz-aware kolon için tutarlılık)."""
    return TZ.localize(dt) if dt.tzinfo is None else dt


def _type_tr(t: str) -> str:
    return "giriş" if t == TYPE_IN else "çıkış"


def _edit_detail(ot: str, ow: datetime, on: Optional[str],
                 nt: str, nw: datetime, nn: Optional[str]) -> str:
    """Eski→yeni farkını okunur metne çevir (audit detayı + tarihçe). Zaman dk hassasiyetinde."""
    parts = []
    if ot != nt:
        parts.append(f"hareket: {_type_tr(ot)}→{_type_tr(nt)}")
    # DB'den okunan datetime UTC tz'li olabilir → her ikisini de Istanbul'a çevir
    ows, nws = ow.astimezone(TZ).strftime("%d.%m %H:%M"), nw.astimezone(TZ).strftime("%d.%m %H:%M")
    if ows != nws:
        parts.append(f"zaman: {ows}→{nws}")
    if (on or "") != (nn or ""):
        parts.append(f"not: '{on or '—'}'→'{nn or '—'}'")
    return "; ".join(parts) if parts else "değişiklik yok"


def _assert_alternation(
    db: Session, personnel_id: int, when: datetime, new_type: str, exclude_id: Optional[int] = None
) -> None:
    """Çift giriş/çıkış engeli: `when`'in zaman-komşuları aynı tip olamaz.

    Hem elle oluşturma hem düzenlemede kullanılır. exclude_id ile düzenlenen kaydın
    kendisi komşu sayılmaz. Geriye-tarihli kayıtlarda da doğru (önceki + sonraki bakılır).
    """
    type_tr = "giriş" if new_type == TYPE_IN else "çıkış"
    other_tr = "çıkış" if new_type == TYPE_IN else "giriş"
    base = db.query(AttendanceLog).filter(
        AttendanceLog.personnel_id == personnel_id, AttendanceLog.deleted_at.is_(None)
    )
    if exclude_id:
        base = base.filter(AttendanceLog.id != exclude_id)
    prev = base.filter(AttendanceLog.punched_at <= when).order_by(desc(AttendanceLog.punched_at)).first()
    nxt = base.filter(AttendanceLog.punched_at > when).order_by(AttendanceLog.punched_at).first()
    if (prev and prev.type == new_type) or (nxt and nxt.type == new_type):
        raise HTTPException(
            status_code=400,
            detail=f"Bu personelin komşu hareketi zaten '{type_tr}'. Art arda çift {type_tr} "
                   f"kaydedilemez — araya '{other_tr}' gerekir.",
        )


def _today_summary(db: Session, personnel_id: int) -> dict:
    """Bugünkü toplam içeride-süre (dakika) + şu an içeride mi."""
    today = datetime.now(TZ).date()
    start = TZ.localize(datetime.combine(today, datetime.min.time()))
    logs = (
        db.query(AttendanceLog)
        .filter(AttendanceLog.personnel_id == personnel_id, AttendanceLog.punched_at >= start,
                AttendanceLog.deleted_at.is_(None))
        .order_by(AttendanceLog.punched_at)
        .all()
    )
    total = 0.0
    open_in: Optional[datetime] = None
    for lg in logs:
        if lg.type == TYPE_IN:
            open_in = lg.punched_at
        elif lg.type == TYPE_OUT and open_in:
            total += (lg.punched_at - open_in).total_seconds()
            open_in = None
    inside = open_in is not None
    if inside:
        total += (datetime.now(TZ) - open_in).total_seconds()
    return {"minutes_today": round(total / 60), "inside": inside}


def _personnel_dict(p: Personnel) -> dict:
    return {
        "id": p.id, "full_name": p.full_name, "employee_code": p.employee_code,
        "department": p.department, "title": p.title, "phone": p.phone, "is_active": p.is_active,
        "created_at": p.created_at.isoformat() if p.created_at else None,
    }


def _norm_hdr(s) -> str:
    """Başlık adını normalize et (TR karakter + boşluk + küçük harf) — kolon eşleştirme için."""
    t = str(s or "").strip().lower()
    for a, b in (("ı", "i"), ("ö", "o"), ("ü", "u"), ("ç", "c"), ("ş", "s"), ("ğ", "g"), ("İ", "i")):
        t = t.replace(a, b)
    return " ".join(t.split())


def _parse_personnel_excel(content: bytes, ext: str) -> List[dict]:
    """Sicil Excel'ini (Sicil No / Ad Soyad / Departman / Görev) satırlara çevir.

    Başlık satırı otomatik bulunur (içinde 'sicil' geçen). xls (xlrd) ve xlsx (openpyxl).
    """
    matrix: List[list] = []
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            matrix.append([ws.cell(r, c).value for c in range(ws.ncols)])
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            matrix.append(list(row))
        wb.close()

    cols: dict = {}
    hdr_idx = None
    for i, row in enumerate(matrix[:25]):
        norm = [_norm_hdr(c) for c in row]
        if any("sicil" in n for n in norm):
            for j, n in enumerate(norm):
                if not n:
                    continue
                if "sicil" in n:
                    cols["code"] = j
                elif "ad soyad" in n or n in ("ad", "adi", "isim", " adi soyadi"):
                    cols["name"] = j
                elif "departman" in n or "bolum" in n:
                    cols["dept"] = j
                elif "gorev" in n or "unvan" in n or "pozisyon" in n:
                    cols["title"] = j
            hdr_idx = i
            break
    if hdr_idx is None or "code" not in cols or "name" not in cols:
        return []

    def _cell(row, key):
        j = cols.get(key)
        return row[j] if j is not None and j < len(row) else None

    out: List[dict] = []
    for row in matrix[hdr_idx + 1:]:
        raw_code = _cell(row, "code")
        raw_name = _cell(row, "name")
        if raw_code is None or not str(raw_name or "").strip():
            continue
        code = str(int(raw_code)) if isinstance(raw_code, float) else str(raw_code).strip()
        if not code:
            continue
        dept = str(_cell(row, "dept")).strip() if _cell(row, "dept") else None
        title = str(_cell(row, "title")).strip() if _cell(row, "title") else None
        out.append({
            "employee_code": code,
            "full_name": str(raw_name).strip(),
            "department": dept or None,
            "title": title or None,
        })
    return out


def _svg_qr(data: str) -> Response:
    qr = segno.make(data, error="m")
    buf = io.BytesIO()
    qr.save(buf, kind="svg", scale=10, border=2)
    return Response(
        content=buf.getvalue(),
        media_type="image/svg+xml",
        headers={"Cache-Control": "no-store, max-age=0"},
    )


# ─── Şemalar ─────────────────────────────────────────────

class PersonnelCreate(BaseModel):
    full_name: str
    employee_code: str
    department: Optional[str] = None
    title: Optional[str] = None
    phone: Optional[str] = None


class PersonnelUpdate(BaseModel):
    full_name: Optional[str] = None
    employee_code: Optional[str] = None
    department: Optional[str] = None
    title: Optional[str] = None
    phone: Optional[str] = None
    is_active: Optional[bool] = None


class SetupRequest(BaseModel):
    token: str


class PunchRequest(BaseModel):
    k: str


class ManualPunch(BaseModel):
    personnel_id: int
    type: str
    punched_at: Optional[datetime] = None
    note: Optional[str] = None


class LogUpdate(BaseModel):
    """Mevcut giriş/çıkış kaydını elle düzenleme (tip / zaman / not)."""
    type: Optional[str] = None
    punched_at: Optional[datetime] = None
    note: Optional[str] = None


class SettingsUpdate(BaseModel):
    refresh_sec: int


# ═══ PUBLIC — Kiosk ekranı ═══════════════════════════════

@router.get("/attendance/kiosk/qr")
def kiosk_qr(key: str = Query(...)):
    """Girişteki ekranın gösterdiği dönen QR (SVG). KIOSK_KEY gerektirir."""
    if not hmac.compare_digest(key, KIOSK_KEY):
        raise HTTPException(status_code=403, detail="Geçersiz kiosk anahtarı")
    url = f"{PUBLIC_BASE}/devam?k={_make_token()}"
    return _svg_qr(url)


@router.get("/attendance/kiosk/config")
def kiosk_config(key: str = Query(...), db: Session = Depends(get_db)):
    """Kiosk ekranının yenileme süresi (saniye). KIOSK_KEY gerektirir."""
    if not hmac.compare_digest(key, KIOSK_KEY):
        raise HTTPException(status_code=403, detail="Geçersiz kiosk anahtarı")
    refresh = _get_refresh(db)
    return {"refresh_sec": refresh, "ttl_sec": _ttl_for(refresh)}


@router.get("/attendance/kiosk/recent")
def kiosk_recent(
    key: str = Query(...),
    limit: int = Query(8, ge=1, le=20),
    db: Session = Depends(get_db),
):
    """Girişteki ekranın sağ paneli için son giriş/çıkış hareketleri. KIOSK_KEY gerektirir."""
    if not hmac.compare_digest(key, KIOSK_KEY):
        raise HTTPException(status_code=403, detail="Geçersiz kiosk anahtarı")
    rows = (
        db.query(AttendanceLog, Personnel)
        .join(Personnel, Personnel.id == AttendanceLog.personnel_id)
        .filter(AttendanceLog.deleted_at.is_(None))
        .order_by(desc(AttendanceLog.punched_at))
        .limit(limit)
        .all()
    )
    return {"items": [{
        "id": lg.id,
        "full_name": p.full_name,
        "department": p.department,
        "type": lg.type,
        "punched_at": lg.punched_at.isoformat(),
    } for lg, p in rows]}


# ═══ PUBLIC — Personel kimlik + basış ════════════════════

@router.post("/attendance/setup")
def setup(data: SetupRequest, request: Request, response: Response, db: Session = Depends(get_db)):
    """Kişisel kurulum linki → kimlik çerezi (pdks_token) set et."""
    punch_limiter.check(f"pdks-setup-{get_client_ip(request)}")
    p = db.query(Personnel).filter(
        Personnel.access_token == data.token.strip(), Personnel.is_active.is_(True)
    ).first()
    if not p:
        raise HTTPException(status_code=404, detail="Geçersiz veya pasif personel linki")
    _set_cookie(response, p.access_token)
    return {"ok": True, "full_name": p.full_name, "employee_code": p.employee_code}


@router.get("/attendance/me")
def me(request: Request, db: Session = Depends(get_db)):
    """Çerezdeki personelin bilgisi + bugünkü durumu."""
    p = _personnel_from_request(request, db)
    if not p:
        raise HTTPException(status_code=401, detail="Personel tanımlı değil — kurulum linkini açın")
    summary = _today_summary(db, p.id)
    last = _last_log(db, p.id)
    return {
        "full_name": p.full_name, "employee_code": p.employee_code, "department": p.department,
        "inside": summary["inside"], "minutes_today": summary["minutes_today"],
        "last_punch": last.punched_at.isoformat() if last else None,
        "last_type": last.type if last else None,
    }


@router.post("/attendance/punch")
def punch(data: PunchRequest, request: Request, db: Session = Depends(get_db)):
    """Kiosk QR'ı okutunca çağrılır — token doğrula, giriş/çıkış kaydet."""
    punch_limiter.check(f"pdks-punch-{get_client_ip(request)}")
    p = _personnel_from_request(request, db)
    if not p:
        raise HTTPException(status_code=401, detail="Personel tanımlı değil — kurulum linkini açın")
    if not _valid_token(data.k, _ttl_for(_get_refresh(db))):
        raise HTTPException(status_code=400, detail="Karekod süresi doldu — ekrandaki güncel kodu tekrar okutun")

    now = datetime.now(TZ)
    last = _last_log(db, p.id)
    if last and (now - last.punched_at).total_seconds() < PUNCH_DEBOUNCE_SEC:
        raise HTTPException(status_code=429, detail="Çok hızlı — birkaç saniye sonra tekrar deneyin")

    new_type = TYPE_OUT if (last and last.type == TYPE_IN) else TYPE_IN
    lg = AttendanceLog(personnel_id=p.id, type=new_type, source=SOURCE_PHONE)
    db.add(lg)
    db.commit()
    # Canlı pano: bağlı yöneticilere sinyal (PII yok — veri izin-korumalı uçtan çekilir)
    manager.send_to_all_sync({"type": WSEvent.ATTENDANCE_UPDATED, "action": "punch"})

    summary = _today_summary(db, p.id)
    return {
        "ok": True,
        "type": new_type,
        "full_name": p.full_name,
        "time": now.strftime("%H:%M"),
        "minutes_today": summary["minutes_today"],
        "message": f"{'Giriş' if new_type == TYPE_IN else 'Çıkış'} yapıldı — hoş geldin {p.full_name.split()[0]}!"
        if new_type == TYPE_IN else f"Çıkış yapıldı — iyi günler {p.full_name.split()[0]}!",
    }


# ═══ YÖNETİCİ — Personel yönetimi ════════════════════════

@router.get("/attendance/personnel")
def list_personnel(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("hr.attendance", "view")),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    include_inactive: bool = Query(True),
):
    q = db.query(Personnel)
    if not include_inactive:
        q = q.filter(Personnel.is_active.is_(True))
    if search:
        like = f"%{search.strip()}%"
        q = q.filter((Personnel.full_name.ilike(like)) | (Personnel.employee_code.ilike(like)))
    total = q.count()
    items = q.order_by(Personnel.full_name).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [_personnel_dict(p) for p in items],
        "total": total, "page": page, "page_size": page_size,
        "pages": math.ceil(total / page_size) if total else 1,
    }


@router.post("/attendance/personnel", status_code=201)
def create_personnel(
    data: PersonnelCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    code = data.employee_code.strip()
    if db.query(Personnel).filter(Personnel.employee_code == code).first():
        raise HTTPException(status_code=400, detail="Bu sicil no zaten kayıtlı")
    p = Personnel(
        full_name=data.full_name.strip(),
        employee_code=code,
        department=(data.department or "").strip() or None,
        title=(data.title or "").strip() or None,
        phone=(data.phone or "").strip() or None,
        access_token=secrets.token_urlsafe(24),
    )
    db.add(p)
    db.flush()
    log_action(db, current_user.id, "create", "personnel", p.id,
               f"Personel: {p.full_name} ({p.employee_code})", get_client_ip(request))
    db.commit()
    db.refresh(p)
    return _personnel_dict(p)


@router.post("/attendance/personnel/import")
async def import_personnel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    """Excel sicil listesinden personel içe aktar — upsert (sicil no = employee_code).

    Beklenen başlıklar (sırası önemsiz): Sicil No, Ad Soyad, Departman, Görev.
    Var olan sicil güncellenir (ad/departman/görev), yoksa yeni eklenir (kişisel token üretilir).
    """
    content = await validate_upload_file(file, allowed_types=["excel"])
    ext = os.path.splitext(file.filename or "")[1].lower() or ".xlsx"
    try:
        rows = _parse_personnel_excel(content, ext)
    except Exception:
        raise HTTPException(status_code=400, detail="Dosya ayrıştırılamadı. Geçerli bir Excel (.xls/.xlsx) yükleyin.")
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="Beklenen kolonlar bulunamadı. Başlıkta en az 'Sicil No' ve 'Ad Soyad' olmalı (Departman/Görev opsiyonel).",
        )

    created = updated = 0
    seen: set = set()
    for r in rows:
        if r["employee_code"] in seen:
            continue
        seen.add(r["employee_code"])
        existing = db.query(Personnel).filter(Personnel.employee_code == r["employee_code"]).first()
        if existing:
            existing.full_name = r["full_name"]
            existing.department = r["department"]
            existing.title = r["title"]
            existing.is_active = True
            updated += 1
        else:
            db.add(Personnel(
                full_name=r["full_name"],
                employee_code=r["employee_code"],
                department=r["department"],
                title=r["title"],
                access_token=secrets.token_urlsafe(24),
            ))
            created += 1
    log_action(db, current_user.id, "import", "personnel", None,
               f"Sicil içe aktarma: {created} yeni, {updated} güncellendi ({len(seen)} satır)",
               get_client_ip(request))
    db.commit()
    return {"ok": True, "created": created, "updated": updated, "total": len(seen)}


@router.patch("/attendance/personnel/{pid}")
def update_personnel(
    pid: int,
    data: PersonnelUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    p = db.query(Personnel).filter(Personnel.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    payload = data.model_dump(exclude_unset=True)
    if "employee_code" in payload and payload["employee_code"]:
        code = payload["employee_code"].strip()
        clash = db.query(Personnel).filter(Personnel.employee_code == code, Personnel.id != pid).first()
        if clash:
            raise HTTPException(status_code=400, detail="Bu sicil no başka personelde")
        p.employee_code = code
    for f in ("full_name", "department", "title", "phone", "is_active"):
        if f in payload:
            setattr(p, f, payload[f])
    log_action(db, current_user.id, "update", "personnel", p.id,
               f"Personel güncellendi: {p.full_name}", get_client_ip(request))
    db.commit()
    db.refresh(p)
    return _personnel_dict(p)


@router.delete("/attendance/personnel/{pid}")
def delete_personnel(
    pid: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    p = db.query(Personnel).filter(Personnel.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    log_action(db, current_user.id, "delete", "personnel", p.id,
               f"Personel silindi: {p.full_name} ({p.employee_code})", get_client_ip(request))
    db.delete(p)  # CASCADE logları siler
    db.commit()
    return {"detail": "Personel silindi"}


@router.get("/attendance/personnel/{pid}/qr")
def personnel_setup_qr(
    pid: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("hr.attendance", "view")),
):
    """Personelin kişisel kurulum linkinin QR'ı (kart basmak için)."""
    p = db.query(Personnel).filter(Personnel.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")
    return _svg_qr(f"{PUBLIC_BASE}/devam/kur?t={p.access_token}")


@router.get("/attendance/personnel/cards.pdf")
def personnel_cards_pdf(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("hr.attendance", "view")),
):
    """Tüm aktif personelin QR kartlarını tek PDF'te üretir (yazdırılıp kesilebilir)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas as rl_canvas

    from app.utils.pdf_fonts import register_turkish_fonts

    base_font, bold_font = register_turkish_fonts()
    people = (
        db.query(Personnel)
        .filter(Personnel.is_active.is_(True))
        .order_by(Personnel.employee_code)
        .all()
    )

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    cols, rows = 2, 5
    margin = 10 * mm
    gap = 5 * mm
    cw = (W - 2 * margin - (cols - 1) * gap) / cols
    ch = (H - 2 * margin - (rows - 1) * gap) / rows
    per_page = cols * rows
    qsize = 28 * mm

    for idx, p in enumerate(people):
        pos = idx % per_page
        if idx > 0 and pos == 0:
            c.showPage()
        col = pos % cols
        row = pos // cols
        x = margin + col * (cw + gap)
        y = H - margin - (row + 1) * ch - row * gap
        c.setStrokeColorRGB(0.82, 0.82, 0.82)
        c.roundRect(x, y, cw, ch, 6, stroke=1, fill=0)
        # QR (kurulum linki)
        qr = segno.make(f"{PUBLIC_BASE}/devam/kur?t={p.access_token}", error="m")
        qb = io.BytesIO()
        qr.save(qb, kind="png", scale=6, border=1)
        qb.seek(0)
        c.drawImage(ImageReader(qb), x + 4 * mm, y + (ch - qsize) / 2, qsize, qsize)
        # Metin
        tx = x + qsize + 8 * mm
        ty = y + ch - 9 * mm
        c.setFillColorRGB(0, 0, 0)
        c.setFont(bold_font, 10)
        c.drawString(tx, ty, (p.full_name or "")[:24])
        c.setFont(base_font, 8)
        c.drawString(tx, ty - 13, f"Sicil: {p.employee_code}")
        if p.department:
            c.drawString(tx, ty - 25, p.department[:22])
        if p.title:
            c.setFillColorRGB(0.3, 0.3, 0.3)
            c.drawString(tx, ty - 37, p.title[:22])
        c.setFont(base_font, 6.5)
        c.setFillColorRGB(0.55, 0.55, 0.55)
        c.drawString(tx, y + 5 * mm, "Uygulamandan 'Tara' ile okut")

    if not people:
        c.setFont(base_font, 12)
        c.drawString(margin, H / 2, "Aktif personel yok")
    c.showPage()
    c.save()
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=personel-qr-kartlari.pdf"},
    )


# ═══ YÖNETİCİ — İzleme + raporlar ════════════════════════

@router.get("/attendance/kiosk-link")
def kiosk_link(_: User = Depends(require_permission("hr.attendance", "view"))):
    """Giriş ekranı için açılacak link (KIOSK_KEY dahil) — admin cihaza kurar."""
    return {"url": f"{PUBLIC_BASE}/devam/ekran?key={KIOSK_KEY}"}


@router.get("/attendance/settings")
def get_settings(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("hr.attendance", "view")),
):
    """PDKS ayarları — QR yenileme süresi + türetilen güvenlik geçerliliği."""
    refresh = _get_refresh(db)
    return {
        "refresh_sec": refresh,
        "ttl_sec": _ttl_for(refresh),
        "min": MIN_REFRESH_SEC,
        "max": MAX_REFRESH_SEC,
    }


@router.patch("/attendance/settings")
def update_settings(
    data: SettingsUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    """QR yenileme süresini güncelle (saniye). Bu modül onay akışından muaftır (ops/HR)."""
    refresh = data.refresh_sec
    if refresh < MIN_REFRESH_SEC or refresh > MAX_REFRESH_SEC:
        raise HTTPException(
            status_code=400,
            detail=f"Süre {MIN_REFRESH_SEC}-{MAX_REFRESH_SEC} saniye arasında olmalı",
        )
    row = db.query(AttendanceSetting).filter(AttendanceSetting.id == 1).first()
    if not row:
        row = AttendanceSetting(id=1, refresh_sec=refresh)
        db.add(row)
    else:
        row.refresh_sec = refresh
    log_action(db, current_user.id, "update", "attendance_settings", 1,
               f"QR yenileme süresi: {refresh}sn", get_client_ip(request))
    db.commit()
    return {
        "refresh_sec": refresh,
        "ttl_sec": _ttl_for(refresh),
        "min": MIN_REFRESH_SEC,
        "max": MAX_REFRESH_SEC,
    }


@router.get("/attendance/status")
def who_is_inside(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("hr.attendance", "view")),
):
    """Şu an içeride olan personeller (son basışı 'in' olanlar)."""
    # Her personelin son log'u
    sub = (
        db.query(AttendanceLog.personnel_id, func.max(AttendanceLog.punched_at).label("mx"))
        .filter(AttendanceLog.deleted_at.is_(None))
        .group_by(AttendanceLog.personnel_id)
        .subquery()
    )
    rows = (
        db.query(AttendanceLog, Personnel)
        .join(sub, (AttendanceLog.personnel_id == sub.c.personnel_id) & (AttendanceLog.punched_at == sub.c.mx))
        .join(Personnel, Personnel.id == AttendanceLog.personnel_id)
        .filter(AttendanceLog.type == TYPE_IN, AttendanceLog.deleted_at.is_(None))
        .order_by(desc(AttendanceLog.punched_at))
        .all()
    )
    inside = [{
        "personnel_id": p.id, "full_name": p.full_name, "department": p.department,
        "since": lg.punched_at.isoformat(),
    } for lg, p in rows]
    return {"inside_count": len(inside), "inside": inside}


@router.get("/attendance/logs")
def list_logs(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("hr.attendance", "view")),
    personnel_id: Optional[int] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
):
    q = db.query(AttendanceLog, Personnel).join(Personnel, Personnel.id == AttendanceLog.personnel_id)
    if personnel_id:
        q = q.filter(AttendanceLog.personnel_id == personnel_id)
    if start_date:
        q = q.filter(AttendanceLog.punched_at >= TZ.localize(datetime.combine(start_date, datetime.min.time())))
    if end_date:
        q = q.filter(AttendanceLog.punched_at < TZ.localize(datetime.combine(end_date + timedelta(days=1), datetime.min.time())))
    total = q.count()
    rows = q.order_by(desc(AttendanceLog.punched_at)).offset((page - 1) * page_size).limit(page_size).all()
    items = [{
        "id": lg.id, "personnel_id": p.id, "full_name": p.full_name, "department": p.department,
        "type": lg.type, "punched_at": lg.punched_at.isoformat(), "source": lg.source,
        "note": lg.note,
        "edited_at": lg.edited_at.isoformat() if lg.edited_at else None,
        "deleted_at": lg.deleted_at.isoformat() if lg.deleted_at else None,
    } for lg, p in rows]
    return {"items": items, "total": total, "page": page, "page_size": page_size,
            "pages": math.ceil(total / page_size) if total else 1}


@router.get("/attendance/summary")
def monthly_summary(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("hr.attendance", "view")),
    month: Optional[str] = Query(None),  # YYYY-MM
):
    """Aylık puantaj — personel başına toplam içeride-süre (dakika) + gün sayısı."""
    now = datetime.now(TZ)
    if month:
        try:
            y, m = map(int, month.split("-"))
        except ValueError:
            raise HTTPException(status_code=400, detail="month formatı YYYY-MM olmalı")
    else:
        y, m = now.year, now.month
    start = TZ.localize(datetime(y, m, 1))
    end = TZ.localize(datetime(y + (m // 12), (m % 12) + 1, 1))

    logs = (
        db.query(AttendanceLog, Personnel)
        .join(Personnel, Personnel.id == AttendanceLog.personnel_id)
        .filter(AttendanceLog.punched_at >= start, AttendanceLog.punched_at < end,
                AttendanceLog.deleted_at.is_(None))
        .order_by(AttendanceLog.personnel_id, AttendanceLog.punched_at)
        .all()
    )
    # Personel başına in→out eşle
    by_p: dict = {}
    for lg, p in logs:
        d = by_p.setdefault(p.id, {"full_name": p.full_name, "department": p.department,
                                    "minutes": 0.0, "days": set(), "open_in": None})
        if lg.type == TYPE_IN:
            d["open_in"] = lg.punched_at
        elif lg.type == TYPE_OUT and d["open_in"]:
            d["minutes"] += (lg.punched_at - d["open_in"]).total_seconds() / 60
            d["days"].add(lg.punched_at.date())
            d["open_in"] = None
    result = [{
        "personnel_id": pid, "full_name": v["full_name"], "department": v["department"],
        "total_minutes": round(v["minutes"]), "total_hours": round(v["minutes"] / 60, 1),
        "days_worked": len(v["days"]),
    } for pid, v in by_p.items()]
    result.sort(key=lambda r: r["full_name"])
    return {"month": f"{y:04d}-{m:02d}", "personnel": result}


@router.post("/attendance/manual", status_code=201)
def manual_punch(
    data: ManualPunch,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    """Yönetici elle giriş/çıkış kaydı (telefonu olmayan / unutulan için).

    - **Durum tutarlılığı:** içerideki kişiye tekrar 'giriş' (veya dışarıdakine 'çıkış')
      engellenir — komşu hareketlerle art arda aynı tip olamaz.
    - **Onay akışı:** hr.attendance için aktif workflow + talep edenin rolü requestor ise
      işlem onaya düşer (202); aksi halde doğrudan kaydedilir.
    """
    if data.type not in (TYPE_IN, TYPE_OUT):
        raise HTTPException(status_code=400, detail="type 'in' veya 'out' olmalı")
    p = db.query(Personnel).filter(Personnel.id == data.personnel_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Personel bulunamadı")

    when = _localize(data.punched_at or datetime.now(TZ))
    _assert_alternation(db, p.id, when, data.type)

    # Onay akışı — aktif workflow + requestor rolü varsa 202 döner, kayıt onaya gider
    payload = data.model_dump()
    payload["punched_at"] = when.isoformat()
    approval_resp = check_approval(db, "hr.attendance", 0, current_user.id, "create", payload)
    if approval_resp:
        return approval_resp

    lg = AttendanceLog(
        personnel_id=p.id, type=data.type, source=SOURCE_MANUAL,
        recorded_by=current_user.id, note=(data.note or "").strip() or None,
        punched_at=when,
    )
    db.add(lg)
    db.flush()
    log_action(db, current_user.id, "manual_punch", "attendance", lg.id,
               f"Elle {data.type} ({when.strftime('%d.%m %H:%M')}): {p.full_name}", get_client_ip(request))
    db.commit()
    manager.send_to_all_sync({"type": WSEvent.ATTENDANCE_UPDATED, "action": "manual"})
    return {"ok": True, "type": data.type, "personnel": p.full_name}


@router.patch("/attendance/logs/{log_id}")
def update_log(
    log_id: int,
    data: LogUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    """Mevcut giriş/çıkış kaydını elle düzenle (tip / zaman / not).

    Çift giriş/çıkış engeli (kendisi hariç komşulara göre) + audit + onay akışına tabi.
    """
    lg = db.query(AttendanceLog).filter(AttendanceLog.id == log_id).first()
    if not lg:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    if lg.deleted_at:
        raise HTTPException(status_code=400, detail="Silinmiş kayıt düzenlenemez")

    fields = data.model_dump(exclude_unset=True)
    if not fields:
        raise HTTPException(status_code=400, detail="Güncellenecek alan yok")
    new_type = fields.get("type") or lg.type
    if new_type not in (TYPE_IN, TYPE_OUT):
        raise HTTPException(status_code=400, detail="type 'in' veya 'out' olmalı")
    new_when = _localize(fields["punched_at"]) if fields.get("punched_at") else lg.punched_at
    _assert_alternation(db, lg.personnel_id, new_when, new_type, exclude_id=lg.id)

    # Onay akışı — payload'da punched_at concrete isoformat
    payload = dict(fields)
    if fields.get("punched_at"):
        payload["punched_at"] = new_when.isoformat()
    approval_resp = check_approval(db, "hr.attendance", lg.id, current_user.id, "update", payload)
    if approval_resp:
        return approval_resp

    old_type, old_when, old_note = lg.type, lg.punched_at, lg.note
    if "type" in fields:
        lg.type = new_type
    if "note" in fields:
        lg.note = (fields["note"] or "").strip() or None
    if fields.get("punched_at"):
        lg.punched_at = new_when
    lg.edited_at = datetime.now(TZ)
    detail = _edit_detail(old_type, old_when, old_note, lg.type, lg.punched_at, lg.note)
    log_action(db, current_user.id, "update", "attendance", lg.id, detail, get_client_ip(request))
    db.commit()
    manager.send_to_all_sync({"type": WSEvent.ATTENDANCE_UPDATED, "action": "edit"})
    return {"ok": True, "id": lg.id, "type": lg.type,
            "punched_at": lg.punched_at.isoformat(), "note": lg.note}


@router.delete("/attendance/logs/{log_id}")
def delete_log(
    log_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    """Giriş/çıkış kaydını sil (soft delete — kayıt kalır, soluk gösterilir). Audit + onay akışına tabi."""
    lg = db.query(AttendanceLog).filter(AttendanceLog.id == log_id).first()
    if not lg:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    if lg.deleted_at:
        raise HTTPException(status_code=400, detail="Kayıt zaten silinmiş")

    approval_resp = check_approval(db, "hr.attendance", lg.id, current_user.id, "delete", {})
    if approval_resp:
        return approval_resp

    log_action(db, current_user.id, "delete", "attendance", lg.id,
               f"Kayıt #{lg.id} silindi ({_type_tr(lg.type)} {lg.punched_at.astimezone(TZ).strftime('%d.%m %H:%M')})",
               get_client_ip(request))
    lg.deleted_at = datetime.now(TZ)  # soft delete
    db.commit()
    manager.send_to_all_sync({"type": WSEvent.ATTENDANCE_UPDATED, "action": "delete"})
    return {"ok": True}


@router.get("/attendance/logs/{log_id}/history")
def log_history(
    log_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("hr.attendance", "view")),
):
    """Bir kaydın değişiklik tarihçesi (audit_logs) + varsa bekleyen onay işlemi."""
    lg = db.query(AttendanceLog).filter(AttendanceLog.id == log_id).first()
    if not lg:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    rows = (
        db.query(AuditLog, User)
        .outerjoin(User, User.id == AuditLog.user_id)
        .filter(AuditLog.entity_type == "attendance", AuditLog.entity_id == log_id)
        .order_by(AuditLog.created_at)
        .all()
    )
    history = [{
        "action": a.action,
        "user_name": (u.full_name if u else None),
        "details": a.details,
        "created_at": a.created_at.isoformat() if a.created_at else None,
    } for a, u in rows]
    pending = (
        db.query(ApprovalRequest)
        .filter(ApprovalRequest.module_code == "hr.attendance",
                ApprovalRequest.entity_id == log_id,
                ApprovalRequest.status == STATUS_PENDING)
        .first()
    )
    return {
        "id": lg.id,
        "edited_at": lg.edited_at.isoformat() if lg.edited_at else None,
        "history": history,
        "pending_action": pending.action_type if pending else None,
    }


# ═══ YÖNETİCİ — Onay bekleyenler (PDKS) ══════════════════

@router.get("/attendance/pending")
def list_pending(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "view")),
):
    """Bekleyen hr.attendance onay talepleri (ekle/düzenle/sil) — pano + filtre + iptal."""
    reqs = (
        db.query(ApprovalRequest)
        .filter(ApprovalRequest.module_code == "hr.attendance",
                ApprovalRequest.status == STATUS_PENDING)
        .order_by(desc(ApprovalRequest.requested_at))
        .all()
    )
    user_ids = {r.requested_by for r in reqs if r.requested_by}
    users = {}
    if user_ids:
        users = {u.id: u.full_name for u in db.query(User).filter(User.id.in_(user_ids)).all()}
    items = []
    for r in reqs:
        try:
            payload = json.loads(r.payload_json) if r.payload_json else {}
        except (json.JSONDecodeError, TypeError):
            payload = {}
        pid = payload.get("personnel_id")
        ptype = payload.get("type")
        ptime = payload.get("punched_at")
        if r.action_type in ("update", "delete") and r.entity_id:
            lg = db.query(AttendanceLog).filter(AttendanceLog.id == r.entity_id).first()
            if lg:
                pid = lg.personnel_id
                ptype = ptype or lg.type
                ptime = ptime or lg.punched_at.isoformat()
        pname = None
        if pid:
            per = db.query(Personnel).filter(Personnel.id == pid).first()
            pname = per.full_name if per else None
        items.append({
            "request_id": r.id,
            "action_type": r.action_type,
            "entity_id": r.entity_id,
            "personnel_id": pid,
            "personnel_name": pname,
            "type": ptype,
            "punched_at": ptime,
            "note": payload.get("note"),
            "requested_by": r.requested_by,
            "requested_by_name": users.get(r.requested_by),
            "requested_at": r.requested_at.isoformat() if r.requested_at else None,
            "can_cancel": r.requested_by == current_user.id,
        })
    return {"items": items, "count": len(items)}


@router.post("/attendance/pending/{request_id}/cancel")
def cancel_pending(
    request_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("hr.attendance", "use")),
):
    """Kendi bekleyen hr.attendance onay talebini iptal et (talep sahibi)."""
    req = db.query(ApprovalRequest).filter(ApprovalRequest.id == request_id).first()
    if not req or req.module_code != "hr.attendance":
        raise HTTPException(status_code=404, detail="Onay talebi bulunamadı")
    if req.requested_by != current_user.id:
        raise HTTPException(status_code=403, detail="Yalnızca kendi talebinizi iptal edebilirsiniz")
    if req.status not in (STATUS_PENDING, STATUS_RETURNED):
        raise HTTPException(status_code=400, detail="Bu talep iptal edilebilir durumda değil")
    process_action(db, req, ACTION_CANCEL, current_user.id, None)
    log_action(db, current_user.id, "cancel", "approval_request", req.id,
               f"PDKS onay talebi iptal edildi ({req.action_type})", get_client_ip(request))
    db.commit()
    manager.send_to_all_sync({"type": WSEvent.APPROVAL_STATUS_CHANGED, "module_code": "hr.attendance"})
    manager.send_to_all_sync({"type": WSEvent.ATTENDANCE_UPDATED, "action": "cancel"})
    return {"ok": True}
