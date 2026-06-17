"""PDKS ortak altyapısı — sabitler, token/cihaz/log yardımcıları, Excel parser, şemalar.

Tüm attendance alt-router'ları (kiosk / personnel / logs) bu modülden `import *` ile besler.
Modül sonundaki `__all__` dunder-olmayan TÜM adı (import edilenler dahil) otomatik dışa aktarır
→ alt-modüllerde isim eksikliği (runtime NameError) olmaz.
"""
import hashlib
import hmac
import io
import json
import math
import os
import secrets
import time
from datetime import date, datetime, timedelta
from typing import List, Optional

import pytz
import segno
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, Response, UploadFile
from pydantic import BaseModel
from sqlalchemy import desc, func
from sqlalchemy.orm import Session

from app.config import settings
from app.constants import WSEvent
from app.database import get_db
from app.middleware.auth import require_permission
from app.middleware.rate_limit import RateLimiter, get_client_ip
from app.models.approval import (
    ACTION_CANCEL,
    STATUS_PENDING,
    STATUS_RETURNED,
    ApprovalRequest,
)
from app.models.attendance_setting import (
    DEFAULT_REFRESH_SEC,
    MAX_REFRESH_SEC,
    MIN_REFRESH_SEC,
    TOKEN_GRACE_SEC,
    AttendanceSetting,
)
from app.models.audit_log import AuditLog
from app.models.personnel import (
    SOURCE_MANUAL,
    SOURCE_PHONE,
    TYPE_IN,
    TYPE_OUT,
    AttendanceLog,
    Personnel,
)
from app.models.user import User
from app.utils.approval_check import check_approval
from app.utils.approval_service import get_pending_approver_ids, process_action
from app.utils.audit import log_action
from app.utils.file_validation import validate_upload_file
from app.websocket.manager import manager

TZ = pytz.timezone("Europe/Istanbul")
SECRET = settings.secret_key.encode()
PUBLIC_BASE = settings.cors_origins.split(",")[0].strip().rstrip("/")
# Kiosk ekranını yetkilendiren stabil, admin-only anahtar (SECRET'ten türetilir)
KIOSK_KEY = hmac.new(SECRET, b"pdks-kiosk-key", hashlib.sha256).hexdigest()[:24]

PUNCH_DEBOUNCE_SEC = 30    # aynı personel bu sürede tekrar basamaz

punch_limiter = RateLimiter(max_requests=40, window_seconds=60)


# ─── Zaman-damgalı token yardımcıları ────────────────────
# Token = "<unix_ts>.<HMAC(SECRET, ts)>" — üretildiği andan itibaren (refresh + GRACE)
# saniye geçerli. Pencere hizalama yok → "geçerlilik süresi" net ve öngörülebilir
# (ekran görüntüsü/bayat QR bu süre dolunca işe yaramaz).

def _sign_ts(ts: int) -> str:
    return hmac.new(SECRET, f"pdks:{ts}".encode(), hashlib.sha256).hexdigest()[:16]


def _make_token() -> str:
    ts = int(time.time())
    return f"{ts}.{_sign_ts(ts)}"


def _valid_token(token: str, ttl_sec: int) -> bool:
    try:
        ts_str, sig = token.split(".", 1)
        ts = int(ts_str)
    except (ValueError, AttributeError):
        return False
    now = int(time.time())
    # +1sn: saat kayması toleransı (ileri); now-ts>TTL: süresi dolmuş
    if ts > now + 1 or now - ts > ttl_sec:
        return False
    return hmac.compare_digest(sig, _sign_ts(ts))


def _get_refresh(db: Session) -> int:
    """Panelden ayarlanan QR yenileme süresi (DB tek satır); yoksa varsayılan."""
    row = db.query(AttendanceSetting).filter(AttendanceSetting.id == 1).first()
    return row.refresh_sec if row else DEFAULT_REFRESH_SEC


def _ttl_for(refresh: int) -> int:
    """Token geçerliliği = yenileme süresi + grace (taze QR'ı tararken pay)."""
    return refresh + TOKEN_GRACE_SEC


def _hash_device(tok: str) -> str:
    return hashlib.sha256(tok.encode()).hexdigest()


def _personnel_from_device(request: Request, db: Session) -> Optional[Personnel]:
    """Basış kimliği = cihaza özel token (X-Pdks-Device başlığı).

    Bu token yalnızca kurulumu yapan telefonun localStorage'ında durur; URL'de/QR'da
    ASLA yer almaz → kişisel link kopyalanıp başka telefonda kullanılsa bile basış yapılamaz
    (anti-buddy-punch). Tek aktif cihaz: hash personnel.device_token_hash ile eşleşmeli.
    """
    tok = request.headers.get("X-Pdks-Device")
    if not tok:
        return None
    return db.query(Personnel).filter(
        Personnel.device_token_hash == _hash_device(tok), Personnel.is_active.is_(True)
    ).first()


def _last_log(db: Session, personnel_id: int) -> Optional[AttendanceLog]:
    return (
        db.query(AttendanceLog)
        .filter(AttendanceLog.personnel_id == personnel_id, AttendanceLog.deleted_at.is_(None))
        .order_by(desc(AttendanceLog.punched_at))
        .first()
    )


def _localize(dt: datetime) -> datetime:
    """Naive datetime'ı Europe/Istanbul'a yerelleştir (tz-aware kolon için tutarlılık)."""
    return TZ.localize(dt) if dt.tzinfo is None else dt


def _type_tr(t: str) -> str:
    return "giriş" if t == TYPE_IN else "çıkış"


def _edit_detail(ot: str, ow: datetime, on: Optional[str],
                 nt: str, nw: datetime, nn: Optional[str]) -> str:
    """Eski→yeni farkını okunur metne çevir (audit detayı + tarihçe). Zaman dk hassasiyetinde."""
    parts = []
    if ot != nt:
        parts.append(f"hareket: {_type_tr(ot)}→{_type_tr(nt)}")
    # DB'den okunan datetime UTC tz'li olabilir → her ikisini de Istanbul'a çevir
    ows, nws = ow.astimezone(TZ).strftime("%d.%m %H:%M"), nw.astimezone(TZ).strftime("%d.%m %H:%M")
    if ows != nws:
        parts.append(f"zaman: {ows}→{nws}")
    if (on or "") != (nn or ""):
        parts.append(f"not: '{on or '—'}'→'{nn or '—'}'")
    return "; ".join(parts) if parts else "değişiklik yok"


def _assert_alternation(
    db: Session, personnel_id: int, when: datetime, new_type: str, exclude_id: Optional[int] = None
) -> None:
    """Çift giriş/çıkış engeli: `when`'in zaman-komşuları aynı tip olamaz.

    Hem elle oluşturma hem düzenlemede kullanılır. exclude_id ile düzenlenen kaydın
    kendisi komşu sayılmaz. Geriye-tarihli kayıtlarda da doğru (önceki + sonraki bakılır).
    """
    type_tr = "giriş" if new_type == TYPE_IN else "çıkış"
    other_tr = "çıkış" if new_type == TYPE_IN else "giriş"
    base = db.query(AttendanceLog).filter(
        AttendanceLog.personnel_id == personnel_id, AttendanceLog.deleted_at.is_(None)
    )
    if exclude_id:
        base = base.filter(AttendanceLog.id != exclude_id)
    prev = base.filter(AttendanceLog.punched_at <= when).order_by(desc(AttendanceLog.punched_at)).first()
    nxt = base.filter(AttendanceLog.punched_at > when).order_by(AttendanceLog.punched_at).first()
    if (prev and prev.type == new_type) or (nxt and nxt.type == new_type):
        raise HTTPException(
            status_code=400,
            detail=f"Bu personelin komşu hareketi zaten '{type_tr}'. Art arda çift {type_tr} "
                   f"kaydedilemez — araya '{other_tr}' gerekir.",
        )


def _today_summary(db: Session, personnel_id: int) -> dict:
    """Bugünkü toplam içeride-süre (dakika) + şu an içeride mi."""
    today = datetime.now(TZ).date()
    start = TZ.localize(datetime.combine(today, datetime.min.time()))
    logs = (
        db.query(AttendanceLog)
        .filter(AttendanceLog.personnel_id == personnel_id, AttendanceLog.punched_at >= start,
                AttendanceLog.deleted_at.is_(None))
        .order_by(AttendanceLog.punched_at)
        .all()
    )
    total = 0.0
    open_in: Optional[datetime] = None
    for lg in logs:
        if lg.type == TYPE_IN:
            open_in = lg.punched_at
        elif lg.type == TYPE_OUT and open_in:
            total += (lg.punched_at - open_in).total_seconds()
            open_in = None
    inside = open_in is not None
    if inside:
        total += (datetime.now(TZ) - open_in).total_seconds()
    return {"minutes_today": round(total / 60), "inside": inside}


def _personnel_dict(p: Personnel) -> dict:
    return {
        "id": p.id, "full_name": p.full_name, "employee_code": p.employee_code,
        "department": p.department, "title": p.title, "phone": p.phone, "is_active": p.is_active,
        "device_bound": bool(p.device_token_hash),
        "device_bound_at": p.device_bound_at.isoformat() if p.device_bound_at else None,
        "created_at": p.created_at.isoformat() if p.created_at else None,
    }


def _norm_hdr(s) -> str:
    """Başlık adını normalize et (TR karakter + boşluk + küçük harf) — kolon eşleştirme için."""
    t = str(s or "").strip().lower()
    for a, b in (("ı", "i"), ("ö", "o"), ("ü", "u"), ("ç", "c"), ("ş", "s"), ("ğ", "g"), ("İ", "i")):
        t = t.replace(a, b)
    return " ".join(t.split())


def _parse_personnel_excel(content: bytes, ext: str) -> List[dict]:
    """Sicil Excel'ini (Sicil No / Ad Soyad / Departman / Görev) satırlara çevir.

    Başlık satırı otomatik bulunur (içinde 'sicil' geçen). xls (xlrd) ve xlsx (openpyxl).
    """
    matrix: List[list] = []
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            matrix.append([ws.cell(r, c).value for c in range(ws.ncols)])
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            matrix.append(list(row))
        wb.close()

    cols: dict = {}
    hdr_idx = None
    for i, row in enumerate(matrix[:25]):
        norm = [_norm_hdr(c) for c in row]
        if any("sicil" in n for n in norm):
            for j, n in enumerate(norm):
                if not n:
                    continue
                if "sicil" in n:
                    cols["code"] = j
                elif "ad soyad" in n or n in ("ad", "adi", "isim", " adi soyadi"):
                    cols["name"] = j
                elif "departman" in n or "bolum" in n:
                    cols["dept"] = j
                elif "gorev" in n or "unvan" in n or "pozisyon" in n:
                    cols["title"] = j
            hdr_idx = i
            break
    if hdr_idx is None or "code" not in cols or "name" not in cols:
        return []

    def _cell(row, key):
        j = cols.get(key)
        return row[j] if j is not None and j < len(row) else None

    out: List[dict] = []
    for row in matrix[hdr_idx + 1:]:
        raw_code = _cell(row, "code")
        raw_name = _cell(row, "name")
        if raw_code is None or not str(raw_name or "").strip():
            continue
        code = str(int(raw_code)) if isinstance(raw_code, float) else str(raw_code).strip()
        if not code:
            continue
        dept = str(_cell(row, "dept")).strip() if _cell(row, "dept") else None
        title = str(_cell(row, "title")).strip() if _cell(row, "title") else None
        out.append({
            "employee_code": code,
            "full_name": str(raw_name).strip(),
            "department": dept or None,
            "title": title or None,
        })
    return out


def _svg_qr(data: str) -> Response:
    qr = segno.make(data, error="m")
    buf = io.BytesIO()
    qr.save(buf, kind="svg", scale=10, border=2)
    return Response(
        content=buf.getvalue(),
        media_type="image/svg+xml",
        headers={"Cache-Control": "no-store, max-age=0"},
    )


# ─── Şemalar ─────────────────────────────────────────────

class PersonnelCreate(BaseModel):
    full_name: str
    employee_code: str
    department: Optional[str] = None
    title: Optional[str] = None
    phone: Optional[str] = None


class PersonnelUpdate(BaseModel):
    full_name: Optional[str] = None
    employee_code: Optional[str] = None
    department: Optional[str] = None
    title: Optional[str] = None
    phone: Optional[str] = None
    is_active: Optional[bool] = None


class SetupRequest(BaseModel):
    token: str


class PunchRequest(BaseModel):
    k: str


class ManualPunch(BaseModel):
    personnel_id: int
    type: str
    punched_at: Optional[datetime] = None
    note: Optional[str] = None


class LogUpdate(BaseModel):
    """Mevcut giriş/çıkış kaydını elle düzenleme (tip / zaman / not)."""
    type: Optional[str] = None
    punched_at: Optional[datetime] = None
    note: Optional[str] = None


class SettingsUpdate(BaseModel):
    refresh_sec: int


# Dunder-olmayan TÜM adı (import/sabit/helper/şema) otomatik dışa aktar → alt-modüllerde
# `from ._helpers import *` ile hiçbir isim eksik kalmaz (eksik-import/NameError riski yok).
__all__ = [_n for _n in dir() if not _n.startswith("__")]
