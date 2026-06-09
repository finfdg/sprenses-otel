"""Ödeme talimat listeleri testleri (finance.cariler izni).

Endpoint'ler: /api/finance/payment-instructions/
"""
import pytest

from app.models.vendor import Vendor

PREFIX = "/api/finance/payment-instructions"


def _make_vendor(db, hesap_kodu="320.99.TEST", hesap_adi="TEST CARI"):
    v = Vendor(hesap_kodu=hesap_kodu, hesap_adi=hesap_adi)
    db.add(v)
    db.flush()
    return v


def _create_list(client, auth_headers, name="Test Talimat", items=None):
    return client.post(
        f"{PREFIX}/",
        json={"name": name, "items": items or []},
        headers=auth_headers,
    )


# ─── Auth ───────────────────────────────────────────────


def test_list_requires_auth(client):
    assert client.get(f"{PREFIX}/").status_code in (401, 403)


def test_create_requires_auth(client):
    assert client.post(f"{PREFIX}/", json={"name": "X"}).status_code in (401, 403)


def test_viewer_cannot_create(client, viewer_user_headers):
    """can_view var, can_use yok → POST 403."""
    res = client.post(f"{PREFIX}/", json={"name": "X"}, headers=viewer_user_headers)
    assert res.status_code == 403


def test_viewer_can_list(client, viewer_user_headers):
    res = client.get(f"{PREFIX}/", headers=viewer_user_headers)
    assert res.status_code == 200


# ─── Liste CRUD ─────────────────────────────────────────


def test_create_empty_list(client, auth_headers):
    res = _create_list(client, auth_headers, name="Boş Liste")
    assert res.status_code == 201, res.text
    data = res.json()
    assert data["name"] == "Boş Liste"
    assert data["item_count"] == 0
    assert data["total_amount"] == 0
    assert data["status"] == "draft"


def test_create_with_items(client, auth_headers):
    res = _create_list(client, auth_headers, items=[
        {"hesap_kodu": "320.1", "hesap_adi": "A Cari", "amount": 1000.50, "balance_snapshot": -1000.50},
        {"hesap_kodu": "320.2", "hesap_adi": "B Cari", "amount": 2500, "balance_snapshot": -2500},
    ])
    assert res.status_code == 201
    data = res.json()
    assert data["item_count"] == 2
    assert data["total_amount"] == 3500.50
    assert len(data["items"]) == 2


def test_get_list_detail(client, auth_headers):
    lid = _create_list(client, auth_headers, items=[
        {"hesap_adi": "X", "amount": 100},
    ]).json()["id"]
    res = client.get(f"{PREFIX}/{lid}", headers=auth_headers)
    assert res.status_code == 200
    assert res.json()["items"][0]["hesap_adi"] == "X"


def test_get_not_found(client, auth_headers):
    assert client.get(f"{PREFIX}/999999", headers=auth_headers).status_code == 404


def test_list_overview(client, auth_headers):
    _create_list(client, auth_headers, name="L1")
    _create_list(client, auth_headers, name="L2")
    res = client.get(f"{PREFIX}/", headers=auth_headers)
    assert res.status_code == 200
    names = [l["name"] for l in res.json()]
    assert "L1" in names and "L2" in names


def test_update_list_name(client, auth_headers):
    lid = _create_list(client, auth_headers, name="Eski").json()["id"]
    res = client.patch(f"{PREFIX}/{lid}", json={"name": "Yeni"}, headers=auth_headers)
    assert res.status_code == 200
    assert res.json()["name"] == "Yeni"


def test_delete_list(client, auth_headers, db):
    lid = _create_list(client, auth_headers).json()["id"]
    assert client.delete(f"{PREFIX}/{lid}", headers=auth_headers).status_code == 204
    assert client.get(f"{PREFIX}/{lid}", headers=auth_headers).status_code == 404


# ─── Kalem yönetimi ─────────────────────────────────────


def test_add_items(client, auth_headers, db):
    v = _make_vendor(db)
    lid = _create_list(client, auth_headers).json()["id"]
    res = client.post(
        f"{PREFIX}/{lid}/items",
        json={"items": [{"vendor_id": v.id, "hesap_kodu": v.hesap_kodu, "hesap_adi": v.hesap_adi, "amount": 5000, "balance_snapshot": -5000}]},
        headers=auth_headers,
    )
    assert res.status_code == 200, res.text
    data = res.json()
    assert data["added"] == 1
    assert data["item_count"] == 1
    assert data["total_amount"] == 5000


# ─── Yapı Kredi toplu ödeme export'u ────────────────────


def test_ykb_export_format(client, auth_headers, db):
    """YKB Excel: sayfa 'ykb excel', tam başlıklar, IBAN boşluksuz, BORÇLU HESAP + tutar + TL."""
    import io

    import openpyxl

    from app.models.payment_instruction import PaymentInstructionItem

    lid = _create_list(client, auth_headers, items=[
        {"hesap_kodu": "320.1", "hesap_adi": "A Cari", "amount": 1000.50, "balance_snapshot": -1000.50},
    ]).json()["id"]
    # IBAN'ı doğrudan set et (boşluklu) — export boşluksuz üretmeli
    it = db.query(PaymentInstructionItem).filter_by(list_id=lid).first()
    it.iban = "TR12 3456 7890 1234 5678 9012 34"
    it.notes = "Fatura"
    db.flush()

    res = client.get(f"{PREFIX}/{lid}/export/ykb-excel?debtor_account=65610029", headers=auth_headers)
    assert res.status_code == 200, res.text
    assert "spreadsheet" in res.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws.title == "ykb excel"
    hdr = [ws.cell(1, c).value for c in range(1, 12)]
    assert hdr[0] == "İŞLEM TARİHİ" and hdr[1] == "BORÇLU HESAP"
    assert hdr[4] == "ALICI HESAP/IBAN" and hdr[5] == "ALICI ADI" and hdr[6] == "TUTAR" and hdr[7] == "DÖVİZ"
    # Veri satırı (R2)
    assert ws.cell(2, 2).value == "65610029"                          # BORÇLU HESAP
    assert ws.cell(2, 5).value == "TR123456789012345678901234"        # IBAN boşluksuz
    assert ws.cell(2, 6).value == "A Cari"                            # ALICI ADI
    assert float(ws.cell(2, 7).value) == 1000.50                     # TUTAR (düz ondalık)
    assert ws.cell(2, 8).value == "TL"                               # DÖVİZ
    assert ws.cell(2, 9).value == "Fatura"                           # AÇIKLAMA


def test_ykb_export_requires_auth(client):
    assert client.get(f"{PREFIX}/1/export/ykb-excel").status_code in (401, 403)


def test_ykb_export_default_debtor_from_account(client, auth_headers, db):
    """debtor_account boşsa BORÇLU HESAP kayıtlı YKB TL hesabının no'sundan gelir."""
    import io
    import uuid

    import openpyxl

    from app.models.bank_account import BankAccount

    db.add(BankAccount(bank_name="Yapı Kredi", currency="TRY", account_no="72821701",
                       iban="TR" + uuid.uuid4().hex[:24].upper(), is_active=True))
    db.flush()
    lid = _create_list(client, auth_headers, items=[{"hesap_adi": "C", "amount": 100}]).json()["id"]
    res = client.get(f"{PREFIX}/{lid}/export/ykb-excel", headers=auth_headers)  # debtor_account YOK
    assert res.status_code == 200, res.text
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert wb.active.cell(2, 2).value == "72821701"  # fallback hesap no


def test_create_with_nonexistent_vendor_id_no_500(client, auth_headers):
    """REGRESYON: var olmayan vendor_id ile liste oluşturma 500 (FK ihlali) vermemeli;
    kalem vendor_id=None ile (snapshot korunarak) eklenmeli. (Üretimde görülen hata.)"""
    res = _create_list(client, auth_headers, items=[
        {"vendor_id": 9999999, "hesap_kodu": "320.X", "hesap_adi": "Silinmiş Cari", "amount": 1500},
    ])
    assert res.status_code == 201, res.text
    data = res.json()
    assert data["item_count"] == 1
    assert data["items"][0]["vendor_id"] is None
    assert data["items"][0]["hesap_adi"] == "Silinmiş Cari"


def test_add_items_nonexistent_vendor_id_no_500(client, auth_headers):
    """REGRESYON: var olmayan vendor_id ile kalem ekleme 500 vermemeli — vendor_id None'a düşer."""
    lid = _create_list(client, auth_headers).json()["id"]
    res = client.post(
        f"{PREFIX}/{lid}/items",
        json={"items": [{"vendor_id": 9999999, "hesap_kodu": "320.Y", "hesap_adi": "Geçersiz", "amount": 5000}]},
        headers=auth_headers,
    )
    assert res.status_code == 200, res.text
    data = res.json()
    assert data["added"] == 1
    assert data["items"][0]["vendor_id"] is None
    assert data["items"][0]["hesap_adi"] == "Geçersiz"


def test_add_duplicate_vendor_skipped(client, auth_headers, db):
    """Aynı vendor_id iki kez eklenmez."""
    v = _make_vendor(db)
    lid = _create_list(client, auth_headers).json()["id"]
    body = {"items": [
        {"vendor_id": v.id, "hesap_adi": v.hesap_adi, "amount": 100},
        {"vendor_id": v.id, "hesap_adi": v.hesap_adi, "amount": 200},
    ]}
    res = client.post(f"{PREFIX}/{lid}/items", json=body, headers=auth_headers)
    assert res.status_code == 200
    data = res.json()
    assert data["added"] == 1
    assert data["skipped"] == 1
    assert data["item_count"] == 1


def test_update_item_amount(client, auth_headers):
    lid = _create_list(client, auth_headers, items=[{"hesap_adi": "X", "amount": 100}]).json()["id"]
    item_id = client.get(f"{PREFIX}/{lid}", headers=auth_headers).json()["items"][0]["id"]
    res = client.patch(f"{PREFIX}/{lid}/items/{item_id}", json={"amount": 999.99}, headers=auth_headers)
    assert res.status_code == 200
    assert res.json()["amount"] == 999.99
    # Toplam güncellendi mi?
    assert client.get(f"{PREFIX}/{lid}", headers=auth_headers).json()["total_amount"] == 999.99


def test_delete_item(client, auth_headers):
    lid = _create_list(client, auth_headers, items=[
        {"hesap_adi": "X", "amount": 100},
        {"hesap_adi": "Y", "amount": 200},
    ]).json()["id"]
    items = client.get(f"{PREFIX}/{lid}", headers=auth_headers).json()["items"]
    res = client.delete(f"{PREFIX}/{lid}/items/{items[0]['id']}", headers=auth_headers)
    assert res.status_code == 204
    after = client.get(f"{PREFIX}/{lid}", headers=auth_headers).json()
    assert after["item_count"] == 1
    assert after["total_amount"] == 200


def test_item_not_found(client, auth_headers):
    lid = _create_list(client, auth_headers).json()["id"]
    assert client.patch(f"{PREFIX}/{lid}/items/999999", json={"amount": 1}, headers=auth_headers).status_code == 404


# ─── Export ─────────────────────────────────────────────


def test_export_excel(client, auth_headers):
    lid = _create_list(client, auth_headers, items=[{"hesap_adi": "X", "amount": 1234.56}]).json()["id"]
    res = client.get(f"{PREFIX}/{lid}/export/excel", headers=auth_headers)
    assert res.status_code == 200
    assert "spreadsheet" in res.headers["content-type"]
    assert res.content[:2] == b"PK"  # xlsx = zip


def test_export_pdf(client, auth_headers):
    lid = _create_list(client, auth_headers, items=[{"hesap_adi": "X", "amount": 1234.56}]).json()["id"]
    res = client.get(f"{PREFIX}/{lid}/export/pdf", headers=auth_headers)
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/pdf"
    assert res.content[:4] == b"%PDF"


def test_export_not_found(client, auth_headers):
    assert client.get(f"{PREFIX}/999999/export/excel", headers=auth_headers).status_code == 404
